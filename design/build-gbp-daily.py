"""Build the daily GBP artwork template and planner sheet from one JSON source.

    python3 design/build-gbp-daily.py

strategy/gbp-daily-content.json is the only file to hand edit. This script
regenerates design/gbp-daily.html from it and rewrites the GBP Daily tab in
strategy/ig-content-strategy.xlsx, keeping whatever Status and Notes are
already typed into that tab. Render the artwork afterwards with

    python3 design/render.py gbp-daily
"""
import json
import os
import datetime

ROOT = os.path.dirname(os.path.dirname(os.path.abspath(__file__)))
CONTENT = os.path.join(ROOT, 'strategy', 'gbp-daily-content.json')
TEMPLATE = os.path.join(ROOT, 'design', 'gbp-daily.html')
SHEET = os.path.join(ROOT, 'strategy', 'ig-content-strategy.xlsx')
TAB = 'GBP Daily'
IMG_DIR = 'content/gbp'


def load():
    with open(CONTENT, encoding='utf-8') as f:
        return json.load(f)


def image_path(post):
    """Flat, date-named so the folder reads in posting order."""
    return f"{IMG_DIR}/{post['date']}-{post['slug']}.jpg"


# --- artwork template ------------------------------------------------------

PHOTO_DIR = 'photos/gbp'
UNSPLASH = ('https://images.unsplash.com/{file}'
            '?fm=jpg&q=82&w=1600&h=1200&fit=crop&crop=entropy&cs=tinysrgb')


def photo_src(photo):
    """Prefer a downloaded file, fall back to the Unsplash CDN.

    design/fetch-gbp-photos.py pulls the photos local, which is what you want:
    renders stay identical offline and do not depend on a CDN. Until then the
    card points at Unsplash directly, and if that will not load either the
    layout drops the circle and widens the text rather than leaving a hole.
    """
    local = os.path.join(ROOT, 'design', PHOTO_DIR, f"{photo['id']}.jpg")
    if os.path.exists(local):
        return f"{PHOTO_DIR}/{photo['id']}.jpg"
    return UNSPLASH.format(file=photo['file'])


def markup(text):
    """*asterisks* become the gold emphasis span."""
    out, gold = [], False
    for chunk in text.split('*'):
        out.append(f'<span class="hi">{chunk}</span>' if gold and chunk else chunk)
        gold = not gold
    return ''.join(out)


# Headline and script start at these sizes and are shrunk to fit by the
# template itself. Estimating widths from character counts was never accurate
# enough for Playfair at 800 or for Pinyon's swashes, so the browser measures.
HEADLINE_SIZE = 80
SCRIPT_SIZE = 64


CSS = """
  @font-face { font-family:'Playfair Display'; src:url('fonts/PlayfairDisplay.ttf');
               font-weight:400 900; }
  @font-face { font-family:'Pinyon Script'; src:url('fonts/PinyonScript-Regular.ttf'); }
  @font-face { font-family:'Jost'; src:url('fonts/Jost.ttf'); font-weight:100 900; }

  :root{
    --cream:#F3EEE5; --ink:#2E2A23; --gold:#C9A85C;
    --deep:#1F1C17;                 /* card ground, a shade under brand ink so
                                       cream type and the photo both lift off it */
    --muted:rgba(243,238,229,.72);
  }
  *{margin:0;padding:0;box-sizing:border-box;}
  body{background:#888;}

  /* Maps crops the 1200x900 to the centre 900x900, so every piece of type sits
     inside .safe. Only the photo is allowed into the 150px bleed either side. */
  .slide{
    width:1200px;height:900px;position:relative;overflow:hidden;
    background:var(--deep);color:var(--cream);font-family:'Jost',sans-serif;
  }
  .safe{position:absolute;left:150px;top:0;width:900px;height:900px;}

  /* A magazine spread rather than a decorative circle. The panel runs the full
     height from x=620 (inside .safe, so 470px in) and bleeds off the right edge.
     Maps keeps 620-1050, which is 430px of photograph: enough to read as a
     spread instead of a cropped shape. */
  .orb{
    position:absolute;left:470px;top:0;width:730px;height:900px;overflow:hidden;
    background:#26221C;
  }
  .orb img{width:100%;height:100%;object-fit:cover;display:none;}
  .has-photo .orb img{display:block;}
  /* the photo must never fight the type sitting to its left */
  .orb::after{
    content:'';position:absolute;inset:0;
    background:linear-gradient(90deg,rgba(31,28,23,.92) 0%,rgba(31,28,23,.34) 16%,
                                     rgba(31,28,23,.06) 42%,rgba(31,28,23,.28) 100%);
  }
  /* the spread's gutter */
  .gutter{position:absolute;left:445px;top:96px;bottom:96px;width:1px;
          background:linear-gradient(180deg,rgba(201,168,92,0),rgba(201,168,92,.65) 22%,
                                            rgba(201,168,92,.65) 78%,rgba(201,168,92,0));}

  /* Deterministic top, not centred. The lockup is absolutely positioned and
     ends at y=191; centring left the gap to whatever each card's content height
     happened to leave over, so shorter cards rode up into the logo. Starting at
     a fixed 248 clears it on every card regardless of content, and lines the
     masthead and text entry up across the whole month. */
  .col{
    position:absolute;left:14px;top:0;width:406px;height:900px;
    display:flex;flex-direction:column;justify-content:flex-start;
    padding:232px 0 140px;
  }
  /* No photo file: the panel stays as a tonal block so the spread keeps its
     asymmetry, rather than collapsing the layout. */
  .foot{
    position:absolute;left:14px;bottom:62px;width:406px;
  }
  .foot .line{height:1px;background:rgba(243,238,229,.22);margin-bottom:16px;}
  .foot .meta{
    display:flex;justify-content:space-between;font-size:13px;font-weight:300;
    letter-spacing:.18em;text-transform:uppercase;color:rgba(243,238,229,.55);
  }

  /* The full Cleo R lockup, not the bare LB mark. logo-original.png is flat
     #3E3E3E with alpha-only anti-aliasing, so brightness(0) invert(1) gives
     clean cream with no halo. The gold files cannot do this: they carry a
     full-canvas alpha 1-31 haze that lights up as a pale rectangle. */
  .lockup{position:absolute;top:64px;left:14px;}
  .service{
    font-size:21px;font-weight:500;letter-spacing:.2em;text-transform:uppercase;
    color:var(--gold);line-height:1.4;
  }
  .rule{width:46px;height:2px;background:var(--gold);margin:20px 0 24px;}
  .headline{
    font-family:'Playfair Display',serif;font-weight:800;
    text-transform:uppercase;line-height:1.0;letter-spacing:.005em;
    font-size:80px;
  }
  .headline .hi{color:var(--gold);}
  .script{
    /* Pinyon's p, g and j hang well left of their origin. The indent lets the
       swash fall where the headline's left axis is, optically aligned, without
       crossing the Maps crop line at x=150. */
    font-family:'Pinyon Script',cursive;font-size:64px;line-height:.95;
    color:var(--gold);margin-top:6px;padding-left:22px;white-space:nowrap;
  }
  .body{
    margin-top:24px;font-size:22px;font-weight:300;line-height:1.52;
    letter-spacing:.02em;color:var(--muted);
  }
  .cta{
    margin-top:30px;display:inline-block;align-self:flex-start;
    border:1px solid rgba(243,238,229,.8);padding:16px 30px;text-align:center;
  }
  .cta .go{font-size:17px;font-weight:500;letter-spacing:.24em;text-transform:uppercase;}
  .cta .at{
    margin-top:7px;font-size:13px;font-weight:300;letter-spacing:.14em;
    text-transform:uppercase;color:var(--muted);
  }
"""


def build_html(data):
    """One 1200x900 card per date. Hard left column, photo right, dark ground."""
    by_id = {p['id']: p for p in json.load(
        open(os.path.join(ROOT, 'strategy', 'gbp-photos.json'), encoding='utf-8'))['photos']}
    posts = {}
    for p in data['posts']:
        photo = by_id[p['photo']]
        posts[p['date']] = {
            'service': p.get('service', data['_meta']['service_line']),
            'headline': markup(p['headline']),

            'script': p['script'],
            'body': p['body'],
            'cta': p['button'],
            'src': photo_src(photo),
            'alt': photo['alt'],
        }
    first = data['posts'][0]['date']
    handle = data['_meta']['instagram']
    return f"""<!DOCTYPE html>
<html>
<head>
<meta charset="utf-8">
<title>Google Business Profile — daily 1200x900</title>
<!-- Generated by design/build-gbp-daily.py. Edit strategy/gbp-daily-content.json
     and strategy/gbp-photos.json, not this file. Preview with ?date=2026-09-14 -->
<style>{CSS}</style>
</head>
<body>
<div id="stage"></div>

<script>
const HEADLINE_SIZE = {HEADLINE_SIZE}, SCRIPT_SIZE = {SCRIPT_SIZE};
const POSTS = {json.dumps(posts, indent=2, ensure_ascii=False)};

const P = POSTS[new URLSearchParams(location.search).get('date') || '{first}'];

document.getElementById('stage').innerHTML = `
<div class="slide has-photo" id="slide">
  <div class="safe">
    <div class="orb"><img id="photo" src="${{P.src}}" alt="${{P.alt}}"></div>
    <div class="gutter"></div>
    <img class="lockup" src="brand/logo-original.png" width="190"
         alt="Luxury Beauty by Cleo R"
         style="filter:brightness(0) invert(1);opacity:.92">
    <div class="col">
      <div class="service">${{P.service}}</div>
      <div class="rule"></div>
      <div class="headline">${{P.headline}}</div>
      <div class="script">${{P.script}}</div>
      <div class="body">${{P.body}}</div>
      <div class="cta">
        <div class="go">${{P.cta}}</div>
        <div class="at">{handle}</div>
      </div>
    </div>
    <div class="foot">
      <div class="line"></div>
      <div class="meta"><span>Oakville &middot; Ontario</span><span>{handle}</span></div>
    </div>
  </div>
</div>`;

// Shrink to fit. A long unbreakable word, or one of Pinyon's trailing swashes,
// overflows the column and would run under the photo circle. Measure and step
// down until it fits, which the estimate from character counts could not do.
function fit(sel, start, floor) {{
  const el = document.querySelector(sel);
  if (!el) return;
  for (let size = start; size > floor; size -= 2) {{
    el.style.fontSize = size + 'px';
    if (el.scrollWidth <= el.clientWidth) return;
  }}
}}
fit('.headline', HEADLINE_SIZE, 40);
fit('.script', SCRIPT_SIZE, 34);

// The column now starts at a fixed y so it can never touch the logo, which means
// a long card grows downward into the footer instead. Same remedy as the
// headline: measure the real boxes and step the body down until it clears.
(function fitBlock() {{
  const cta = document.querySelector('.cta');
  const foot = document.querySelector('.foot');
  const body = document.querySelector('.body');
  if (!cta || !foot || !body) return;
  const clears = () => foot.getBoundingClientRect().top
                     - cta.getBoundingClientRect().bottom >= 14;
  for (let size = 22; size >= 16.5 && !clears(); size -= 0.5) {{
    body.style.fontSize = size + 'px';
  }}
  // still tight on the very longest copy: give the script back some room too
  for (let size = parseFloat(getComputedStyle(document.querySelector('.script')).fontSize);
       size >= 34 && !clears(); size -= 2) {{
    document.querySelector('.script').style.fontSize = size + 'px';
  }}
}})();

// no photo available: drop the circle, the column takes the full safe square
document.getElementById('photo').addEventListener('error', () => {{
  document.getElementById('slide').classList.remove('has-photo');
}});
</script>
</body>
</html>
"""


# --- planner tab -----------------------------------------------------------

HEADERS = ['Date', 'Day', 'Source', 'Pillar', 'Post Text (short)', 'Chars',
           'Description (long, SEO)', 'Desc Chars', 'CTA Button', 'CTA Link',
           'Image File', 'Open Folder', 'Open Image', 'Preview', 'Photo Credit',
           'Status', 'Notes']


def build_sheet(data):
    import openpyxl
    from openpyxl.styles import Font, Alignment, PatternFill
    from openpyxl.utils import get_column_letter

    meta = data['_meta']
    wb = openpyxl.load_workbook(SHEET)

    # keep anything already typed into Status / Notes before the tab is replaced
    kept = {}
    if TAB in wb.sheetnames:
        old = wb[TAB]
        head = [c.value for c in old[1]]
        if 'Date' in head:
            di = head.index('Date')
            si = head.index('Status') if 'Status' in head else None
            ni = head.index('Notes') if 'Notes' in head else None
            for row in old.iter_rows(min_row=2, values_only=True):
                key = row[di]
                key = key.strftime('%Y-%m-%d') if hasattr(key, 'strftime') else str(key)
                kept[key] = (row[si] if si is not None else None,
                             row[ni] if ni is not None else None)
        del wb[TAB]

    ws = wb.create_sheet(TAB, wb.sheetnames.index('Posting Schedule') + 1)

    ink = Font(name='Calibri', size=11)
    ws.append(HEADERS)
    for c in ws[1]:
        c.font = Font(bold=True, color='FFFFFF')
        c.fill = PatternFill('solid', fgColor='2E2A23')
        c.alignment = Alignment(vertical='center')
    ws.freeze_panes = 'C2'

    raw = f"https://raw.githubusercontent.com/RepoKing23/InstagramCarousel/{meta['branch']}/"
    # every daily image lives in one flat folder, so this link is the same per row
    folder = (f"https://github.com/RepoKing23/InstagramCarousel/tree/"
              f"{meta['branch']}/{IMG_DIR}")
    by_id = {q['id']: q for q in json.load(
        open(os.path.join(ROOT, 'strategy', 'gbp-photos.json'), encoding='utf-8'))['photos']}

    for p in data['posts']:
        img = image_path(p)
        photo = by_id[p['photo']]
        status, notes = kept.get(p['date'], (None, None))
        ws.append([
            datetime.datetime.strptime(p['date'], '%Y-%m-%d'),
            p['day'], p['source'], p['pillar'], p['text'], len(p['text']),
            p['description'], len(p['description']),
            meta['cta_button'], meta['instagram_url'], img.split('/')[-1],
            'Open folder', 'Open image', f'=IMAGE("{raw}{img}")',
            f"{photo['by']} / Unsplash", status or 'Ready', notes or '',
        ])
        r = ws.max_row
        ws.cell(r, 1).number_format = 'yyyy-mm-dd'
        link = ws.cell(r, HEADERS.index('CTA Link') + 1)
        link.hyperlink = meta['instagram_url']
        link.value = meta['instagram']
        link.style = 'Hyperlink'
        fo = ws.cell(r, HEADERS.index('Open Folder') + 1)
        fo.hyperlink = folder
        fo.style = 'Hyperlink'
        # links straight at the jpg, so it opens ready to save and post
        im = ws.cell(r, HEADERS.index('Open Image') + 1)
        im.hyperlink = f'{raw}{img}'
        im.style = 'Hyperlink'
        cr = ws.cell(r, HEADERS.index('Photo Credit') + 1)
        cr.hyperlink = f"https://unsplash.com/photos/{photo['id']}"
        cr.style = 'Hyperlink'
        for col in ('Post Text (short)', 'Description (long, SEO)', 'Notes'):
            ws.cell(r, HEADERS.index(col) + 1).alignment = Alignment(wrap_text=True, vertical='top')
        # Sunday and Wednesday have no Instagram post behind them
        if p['source'] == 'GBP only':
            for c in ws[r]:
                c.fill = PatternFill('solid', fgColor='F3EEE5')
        for c in ws[r]:
            if not c.font.bold:
                c.font = ink

    widths = {'A': 12, 'B': 6, 'C': 32, 'D': 17, 'E': 54, 'F': 7, 'G': 96, 'H': 8,
              'I': 12, 'J': 26, 'K': 32, 'L': 13, 'M': 13, 'N': 14, 'O': 24,
              'P': 10, 'Q': 28}
    for col, w in widths.items():
        ws.column_dimensions[col].width = w
    for r in range(2, ws.max_row + 1):
        ws.row_dimensions[r].height = 132

    wb.save(SHEET)
    return ws.max_row - 1


def main():
    data = load()
    with open(TEMPLATE, 'w', encoding='utf-8') as f:
        f.write(build_html(data))
    print(f'wrote {os.path.relpath(TEMPLATE, ROOT)}  ({len(data["posts"])} days)')
    n = build_sheet(data)
    print(f'wrote "{TAB}" tab in {os.path.relpath(SHEET, ROOT)}  ({n} rows)')


if __name__ == '__main__':
    main()
