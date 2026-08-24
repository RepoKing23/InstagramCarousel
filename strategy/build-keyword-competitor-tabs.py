#!/usr/bin/env python3
"""
Adds the Target Keywords and Competitors tabs to
strategy/seo-roadmap-tracker.xlsx, and the pointer block on Start Here that
sends you to them.

Run from the repo root:
    pip install openpyxl
    python3 strategy/build-keyword-competitor-tabs.py

Idempotent — it deletes and rebuilds both tabs on every run, so re-run it
after editing the lists below. It never touches the other tabs beyond the
one pointer block at the bottom of Start Here.

The keyword list starts from the forty terms Cleo prioritised herself and
extends them into the money modifiers (cost, near me, best), the treatments
named in her Google description that the forty missed (dermaplaning,
electrolysis, jawline slimming), the surrounding towns, and the questions
people ask before they book.

Volume, difficulty, CPC and current position are deliberately blank. They
come from Semrush against the ca database, and the Semrush account was out
of API units when this was built. Fill the SEMRUSH_PULLED date on Start Here
when the numbers land.

Palette and type match build-seo-tracker.py.
"""

import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.worksheet.datavalidation import DataValidation
from openpyxl.formatting.rule import FormulaRule, ColorScaleRule
from openpyxl.comments import Comment
from openpyxl.utils import get_column_letter

BOOK = "strategy/seo-roadmap-tracker.xlsx"

INK = "2E2A23"
GOLD = "C9A85C"
CREAM_DEEP = "E5DCCA"
CREAM = "F3EEE5"
GREEN = "DDEBD8"
AMBER = "FBEFD4"
RED = "F6DCDA"
GREY = "6E6557"

STATUSES = ["Not started", "Drafting", "Live", "Ranking", "Not applicable"]
PRIORITIES = ["High", "Medium", "Watch"]

# Pages the keyword map points at. The site is not built yet; these are the
# URLs the sitemap in the week-of-24-August planning block should produce.
HOME = "/"
ABOUT = "/about"

# ------------------------------------------------------------------ keywords
# (priority, keyword, cluster, location, intent, target page, page type, note)
# Priorities 1-40 are Cleo's own list, kept in her order and her wording.
SEED = [
    (1,  "Botox Oakville", "Botox", "Oakville", "Transactional", "/botox-oakville", "Service page", "Primary money term. Metrics tab already tracks the map pack for this"),
    (2,  "Dermal Fillers Oakville", "Dermal Fillers", "Oakville", "Transactional", "/dermal-fillers-oakville", "Service page", ""),
    (3,  "Lip Fillers Oakville", "Lip Fillers", "Oakville", "Transactional", "/lip-filler-oakville", "Service page", "Plural. Keep both this and #19 — Google treats them as one, but track the wording that wins"),
    (4,  "Medical Aesthetics Oakville", "Medical Aesthetics", "Oakville", "Commercial", HOME, "Home", ""),
    (5,  "Nurse Practitioner Oakville", "Nurse Practitioner", "Oakville", "Commercial", ABOUT, "About", "Broad — catches medical NP searches too. Watch the traffic quality"),
    (6,  "Cosmetic Nurse Practitioner Oakville", "Nurse Practitioner", "Oakville", "Commercial", ABOUT, "About", "The differentiator. Almost nobody in Oakville is NP-led"),
    (7,  "Skin Boosters Oakville", "Skin Boosters", "Oakville", "Transactional", "/skin-boosters-oakville", "Service page", ""),
    (8,  "Fat Dissolving Oakville", "Fat Dissolve", "Oakville", "Transactional", "/fat-dissolving-oakville", "Service page", ""),
    (9,  "Weight Loss Clinic Oakville", "Weight Loss", "Oakville", "Commercial", "/medical-weight-loss-oakville", "Service page", ""),
    (10, "Medical Weight Loss Oakville", "Weight Loss", "Oakville", "Transactional", "/medical-weight-loss-oakville", "Service page", ""),
    (11, "IV Therapy Oakville", "IV Vitamins", "Oakville", "Transactional", "/iv-therapy-oakville", "Service page", ""),
    (12, "IV Vitamin Therapy Oakville", "IV Vitamins", "Oakville", "Transactional", "/iv-therapy-oakville", "Service page", "Same term as #36 — the duplicate is in the source list, keep one row"),
    (13, "Aesthetic Clinic Oakville", "Medical Aesthetics", "Oakville", "Commercial", HOME, "Home", ""),
    (14, "Cosmetic Injectables Oakville", "Injectables", "Oakville", "Commercial", "/injectables-oakville", "Service page", "Hub page — links down to botox, filler and lip filler"),
    (15, "Weight Loss Injections Oakville", "Weight Loss", "Oakville", "Transactional", "/medical-weight-loss-oakville", "Service page", "Check the copy against CPSO advertising rules before it goes live"),
    (16, "Botox Injections Oakville", "Botox", "Oakville", "Transactional", "/botox-oakville", "Service page", ""),
    (17, "Botox Clinic Oakville", "Botox", "Oakville", "Commercial", "/botox-oakville", "Service page", ""),
    (18, "Botox Treatment Oakville", "Botox", "Oakville", "Transactional", "/botox-oakville", "Service page", ""),
    (19, "Lip Filler Oakville", "Lip Fillers", "Oakville", "Transactional", "/lip-filler-oakville", "Service page", "Singular. The H1 should use whichever of #3 and #19 Semrush gives more volume"),
    (20, "Facial Fillers Oakville", "Dermal Fillers", "Oakville", "Transactional", "/dermal-fillers-oakville", "Service page", ""),
    (21, "Cheek Filler Oakville", "Dermal Fillers", "Oakville", "Transactional", "/dermal-fillers-oakville", "Section on page", ""),
    (22, "Chin Filler Oakville", "Dermal Fillers", "Oakville", "Transactional", "/dermal-fillers-oakville", "Section on page", ""),
    (23, "Jawline Filler Oakville", "Dermal Fillers", "Oakville", "Transactional", "/dermal-fillers-oakville", "Section on page", "Not the same as jawline slimming (#63). Different treatment, different intent"),
    (24, "Skin Booster Treatment Oakville", "Skin Boosters", "Oakville", "Transactional", "/skin-boosters-oakville", "Service page", ""),
    (25, "Skin Rejuvenation Oakville", "Skin Boosters", "Oakville", "Commercial", "/skin-boosters-oakville", "Service page", ""),
    (26, "Fat Dissolving Injections Oakville", "Fat Dissolve", "Oakville", "Transactional", "/fat-dissolving-oakville", "Service page", ""),
    (27, "Fat Dissolve Oakville", "Fat Dissolve", "Oakville", "Transactional", "/fat-dissolving-oakville", "Service page", ""),
    (28, "Body Contouring Oakville", "Fat Dissolve", "Oakville", "Commercial", "/fat-dissolving-oakville", "Service page", "Competitive — the laser and device clinics own this one. Long game"),
    (29, "Non Surgical Fat Reduction Oakville", "Fat Dissolve", "Oakville", "Commercial", "/fat-dissolving-oakville", "Service page", ""),
    (30, "Weight Loss Treatment Oakville", "Weight Loss", "Oakville", "Commercial", "/medical-weight-loss-oakville", "Service page", ""),
    (31, "Medical Weight Management Oakville", "Weight Loss", "Oakville", "Commercial", "/medical-weight-loss-oakville", "Service page", ""),
    (32, "Medically Supervised Weight Loss Oakville", "Weight Loss", "Oakville", "Commercial", "/medical-weight-loss-oakville", "Service page", "The NP angle is the whole argument here"),
    (33, "Weight Management Oakville", "Weight Loss", "Oakville", "Informational", "/medical-weight-loss-oakville", "Service page", ""),
    (34, "IV Hydration Oakville", "IV Vitamins", "Oakville", "Transactional", "/iv-therapy-oakville", "Service page", ""),
    (35, "Vitamin IV Oakville", "IV Vitamins", "Oakville", "Transactional", "/iv-therapy-oakville", "Service page", ""),
    (36, "IV Vitamin Therapy Oakville", "IV Vitamins", "Oakville", "Transactional", "/iv-therapy-oakville", "Service page", "Duplicate of #12 in the source list — worked as one row, priority 12"),
    (37, "Aesthetic Nurse Practitioner Oakville", "Nurse Practitioner", "Oakville", "Commercial", ABOUT, "About", ""),
    (38, "Anti Aging Clinic Oakville", "Anti Aging", "Oakville", "Commercial", HOME, "Home", ""),
    (39, "Injectables Oakville", "Injectables", "Oakville", "Commercial", "/injectables-oakville", "Service page", ""),
    (40, "Cosmetic Injection Clinic Oakville", "Injectables", "Oakville", "Commercial", "/injectables-oakville", "Service page", ""),
]

EXPANSION = [
    # The money modifiers. These convert; the head terms mostly window-shop.
    (41, "botox near me", "Botox", "Oakville", "Transactional", "GBP + /botox-oakville", "GBP / Service page", "Won on the Google profile, not the site. Proximity and reviews decide it"),
    (42, "med spa near me", "Medical Aesthetics", "Oakville", "Transactional", "GBP + " + HOME, "GBP / Home", "Same — this is a map pack term"),
    (43, "lip filler near me", "Lip Fillers", "Oakville", "Transactional", "GBP + /lip-filler-oakville", "GBP / Service page", ""),
    (44, "botox cost oakville", "Botox", "Oakville", "Commercial", "/botox-oakville", "Section on page", "Publish per-unit pricing. Most Oakville clinics hide it — that is the opening"),
    (45, "how much is botox in oakville", "Botox", "Oakville", "Informational", "/botox-oakville", "FAQ", "FAQ schema. Answer in the first sentence"),
    (46, "lip filler cost oakville", "Lip Fillers", "Oakville", "Commercial", "/lip-filler-oakville", "Section on page", ""),
    (47, "dermal filler cost oakville", "Dermal Fillers", "Oakville", "Commercial", "/dermal-fillers-oakville", "Section on page", ""),
    (48, "best botox oakville", "Botox", "Oakville", "Commercial", HOME, "Home", "Review count carries this one. See the Reviews tab"),
    (49, "best lip filler injector oakville", "Lip Fillers", "Oakville", "Commercial", "/lip-filler-oakville", "Service page", ""),
    (50, "best med spa oakville", "Medical Aesthetics", "Oakville", "Commercial", HOME, "Home", ""),
    (51, "lip filler before and after oakville", "Lip Fillers", "Oakville", "Informational", "/lip-filler-oakville", "Gallery", "The before-and-after carousel already in design/ feeds this"),
    (52, "botox before and after oakville", "Botox", "Oakville", "Informational", "/botox-oakville", "Gallery", ""),
    # Treatments named in the Google description that the forty missed.
    (53, "dermaplaning oakville", "Dermaplaning", "Oakville", "Transactional", "/dermaplaning-oakville", "Service page", "In the Google long description but missing from the seed list"),
    (54, "dermaplaning near me", "Dermaplaning", "Oakville", "Transactional", "GBP + /dermaplaning-oakville", "GBP / Service page", ""),
    (55, "electrolysis oakville", "Electrolysis", "Oakville", "Transactional", "/electrolysis-oakville", "Service page", "Same — offered, never targeted. Low volume, very low competition"),
    (56, "permanent hair removal oakville", "Electrolysis", "Oakville", "Commercial", "/electrolysis-oakville", "Service page", "Laser clinics own this. Electrolysis is the honest differentiator"),
    (57, "electrolysis hair removal oakville", "Electrolysis", "Oakville", "Transactional", "/electrolysis-oakville", "Service page", ""),
    (58, "anti wrinkle injections oakville", "Botox", "Oakville", "Transactional", "/botox-oakville", "Service page", "The wording used in the Google description. Keep site and profile consistent"),
    (59, "wrinkle treatment oakville", "Botox", "Oakville", "Commercial", "/botox-oakville", "Service page", ""),
    (60, "forehead lines treatment oakville", "Botox", "Oakville", "Transactional", "/botox-oakville", "Section on page", ""),
    (61, "crows feet treatment oakville", "Botox", "Oakville", "Transactional", "/botox-oakville", "Section on page", ""),
    (62, "masseter botox oakville", "Botox", "Oakville", "Transactional", "/jawline-slimming-oakville", "Service page", "The clinical name for jawline slimming. Higher intent than the plain-English version"),
    (63, "jawline slimming oakville", "Botox", "Oakville", "Transactional", "/jawline-slimming-oakville", "Service page", "In the Google description, missing from the seed list"),
    (64, "teeth grinding botox oakville", "Botox", "Oakville", "Commercial", "/jawline-slimming-oakville", "FAQ", "Confirm the clinic treats bruxism before publishing"),
    (65, "lip flip oakville", "Lip Fillers", "Oakville", "Transactional", "/lip-filler-oakville", "Section on page", "Botox, not filler — but people search it beside lip filler"),
    (66, "under eye filler oakville", "Dermal Fillers", "Oakville", "Transactional", "/dermal-fillers-oakville", "Section on page", ""),
    (67, "tear trough filler oakville", "Dermal Fillers", "Oakville", "Transactional", "/dermal-fillers-oakville", "Section on page", ""),
    (68, "nasolabial fold filler oakville", "Dermal Fillers", "Oakville", "Transactional", "/dermal-fillers-oakville", "Section on page", ""),
    (69, "skin booster cost oakville", "Skin Boosters", "Oakville", "Commercial", "/skin-boosters-oakville", "Section on page", ""),
    (70, "profhilo oakville", "Skin Boosters", "Oakville", "Transactional", "/skin-boosters-oakville", "Section on page", "Brand term — only target it if that product is actually stocked"),
    (71, "weight loss injections cost oakville", "Weight Loss", "Oakville", "Commercial", "/medical-weight-loss-oakville", "Section on page", ""),
    (72, "weight loss clinic near me", "Weight Loss", "Oakville", "Transactional", "GBP + /medical-weight-loss-oakville", "GBP / Service page", ""),
    (73, "iv drip oakville", "IV Vitamins", "Oakville", "Transactional", "/iv-therapy-oakville", "Service page", ""),
    (74, "iv therapy near me", "IV Vitamins", "Oakville", "Transactional", "GBP + /iv-therapy-oakville", "GBP / Service page", ""),
    (75, "vitamin b12 injection oakville", "IV Vitamins", "Oakville", "Transactional", "/iv-therapy-oakville", "Section on page", "Confirm it is offered"),
    (76, "med spa oakville", "Medical Aesthetics", "Oakville", "Commercial", HOME, "Home", "The category term. Matches the Medical Spa primary category on the profile"),
    (77, "medical spa oakville", "Medical Aesthetics", "Oakville", "Commercial", HOME, "Home", ""),
    (78, "nurse injector oakville", "Nurse Practitioner", "Oakville", "Commercial", ABOUT, "About", "Already a tracked map pack term on the Metrics tab"),
    (79, "np injector oakville", "Nurse Practitioner", "Oakville", "Commercial", ABOUT, "About", ""),
    # The rest of the service area named on Start Here.
    (80, "botox burlington", "Botox", "Burlington", "Transactional", "/med-spa-burlington", "Location page", "Only build the location page if there is real copy for it. A thin one hurts"),
    (81, "lip filler burlington", "Lip Fillers", "Burlington", "Transactional", "/med-spa-burlington", "Location page", ""),
    (82, "med spa burlington", "Medical Aesthetics", "Burlington", "Commercial", "/med-spa-burlington", "Location page", ""),
    (83, "botox milton", "Botox", "Milton", "Transactional", "/med-spa-milton", "Location page", ""),
    (84, "med spa milton", "Medical Aesthetics", "Milton", "Commercial", "/med-spa-milton", "Location page", ""),
    (85, "botox mississauga", "Botox", "Mississauga", "Transactional", "/med-spa-mississauga", "Location page", "Big market, heavy competition. Month five or six, not before"),
    (86, "lip filler mississauga", "Lip Fillers", "Mississauga", "Transactional", "/med-spa-mississauga", "Location page", ""),
    (87, "med spa glen abbey", "Medical Aesthetics", "Glen Abbey", "Commercial", HOME, "Home / GBP", "Neighbourhood terms. Low volume, almost no competition, close to the clinic"),
    (88, "botox bronte oakville", "Botox", "Bronte", "Transactional", HOME, "Home / GBP", ""),
    (89, "med spa joshua creek", "Medical Aesthetics", "Joshua Creek", "Commercial", HOME, "Home / GBP", ""),
    (90, "med spa halton", "Medical Aesthetics", "Halton Region", "Commercial", HOME, "Home", ""),
    # Brand. Zero volume today; that is the point of tracking it.
    (91, "luxury beauty by cleo", "Brand", "Oakville", "Navigational", HOME, "Home", "Baseline is near zero. Growth here is the honest measure of the whole six months"),
    (92, "luxury beauty by cleo r oakville", "Brand", "Oakville", "Navigational", HOME, "Home", ""),
    (93, "cleo rukovo np", "Brand", "Oakville", "Navigational", ABOUT, "About", "The practitioner name. Healthcare directories index the person"),
    (94, "cleo rukovo oakville", "Brand", "Oakville", "Navigational", ABOUT, "About", ""),
    # Questions. Blog and FAQ fuel, and Instagram caption fuel.
    (95, "how long does lip filler last", "Lip Fillers", "National", "Informational", "/blog/how-long-does-lip-filler-last", "Blog", "Not local — ranks nationally, feeds the funnel and the carousels"),
    (96, "botox vs filler difference", "Injectables", "National", "Informational", "/blog/botox-vs-filler", "Blog", "Already a carousel in med-spa-carousel-ideas.md. Write once, use twice"),
    (97, "how often should you get botox", "Botox", "National", "Informational", "/blog/how-often-botox", "Blog", ""),
    (98, "what to expect first botox appointment", "Botox", "National", "Informational", "/blog/first-botox-appointment", "Blog", ""),
    (99, "how long does lip filler swelling last", "Lip Fillers", "National", "Informational", "/blog/lip-filler-swelling", "Blog", ""),
    (100, "is botox safe", "Botox", "National", "Informational", "/blog/is-botox-safe", "Blog", "The myths carousel in design/ covers this. Reuse the copy"),
]

KEYWORDS = [k for k in SEED if k[0] != 36] + EXPANSION

# --------------------------------------------------------------- competitors
# Everything here was verified from the clinic's own site or its Google
# listing in August 2026. Semrush columns are blank by design — see the note
# at the top of this file.
# (name, domain, tier, area, what they win on, priority, overlap, note)
COMPETITORS = [
    ("Allure Laser & Skin Studio", "allurelaserskinstudio.ca", "Same address", "3060 Preserve Drive, Oakville",
     "The listing that already sits at your address. Laser, IPL, facials, microneedling — and it lists Cleo as one of its practitioners",
     "High", "Every Oakville term, plus the address itself",
     "Not really a competitor — a NAP collision. Until the suite question is settled, Google may fold the two listings together. This is the single biggest local risk in the file"),
    ("Oakville Flawless Cosmetic", "oakvilleflawlesscosmetic.ca", "Direct — nurse-led", "Oakville",
     "Nurse injector with physician oversight, a decade in. Closest thing to a like-for-like rival",
     "High", "Lip filler, dermal filler, botox, injectables",
     "Study their service page structure first. Same size, same story, further ahead"),
    ("Secret Faces", "secretfaces.com", "Direct — nurse-led", "Oakville",
     "Neighbourhood boutique, RN injector. Botox, Dysport, fillers, PRP, microneedling",
     "High", "Botox, dermal filler, injectables",
     "Boutique positioning, same as yours. Watch how they write about being small"),
    ("Boutique Medical Aesthetics", "boutiquemedicalaesthetics.com", "Direct — nurse-led", "Oakville",
     "RN-founded, forty years of nursing, subtle-results positioning",
     "High", "Botox, filler, medical aesthetics",
     "The experience claim is their whole pitch. Yours is the NP scope of practice"),
    ("the pür glow Medical Spa", "thepurglow.com", "Direct — med spa", "Oakville",
     "Botox, Dysport, Juvederm, Morpheus8, Forma, Lumecca. Ranks for best botox oakville",
     "High", "Botox, filler, med spa oakville",
     "Device-heavy. You will not match the equipment; do not try to"),
    ("Fyxson Medical Aesthetics", "fyxsonmedical.com", "Direct — med spa", "Oakville",
     "Botox, fillers, PRP, microneedling, IV vitamins, laser. Strong service-page SEO",
     "High", "Botox, filler, IV vitamins, med spa oakville",
     "The only local rival also targeting IV therapy. Read their IV page before writing yours"),
    ("Sun & Shade Med Spa", "sunandshademedspa.com", "Direct — med spa", "Oakville",
     "Botox, fillers, Sculptra, PRP, microneedling, laser hair removal, medical facials",
     "Medium", "Botox, filler, med spa oakville", ""),
    ("APT Medical Aesthetics", "aptmedicalaesthetics.com", "Direct — med spa", "Oakville",
     "Botox, fillers, BBL, Halo, Morpheus8. Calls itself the leading Oakville med spa",
     "Medium", "Botox, filler, medical aesthetics oakville", ""),
    ("IMPACT Medical Cosmetics", "impactcosmetic.com", "Direct — med spa", "128 Trafalgar Road, Oakville",
     "Competes on price — Oakville's most affordable med spa",
     "Medium", "Botox, botox cost oakville",
     "The price-led rival. Publishing your own pricing is how you answer this without matching it"),
    ("Distill Laser Clinic", "distilllaserclinic.com", "Direct — laser led", "Oakville",
     "Physician and RN injectors, laser-first clinic with a lip injections page",
     "Medium", "Lip filler, injectables", ""),
    ("Skin Vitality", "skinvitality.ca", "Chain", "Oakville plus multiple Ontario locations",
     "Largest chain in the market. Ten MDs, forty NPs, a page for every treatment in every city",
     "Watch", "Every service term, every city term",
     "Do not chase them on the head terms. They lose on the map pack and on who actually holds the needle"),
    ("ICLS Dermatology & Plastic Surgery", "icls.ca", "Authority", "Oakville",
     "Twenty-five years, dermatology plus plastic surgery plus research. Heavy domain authority",
     "Watch", "Anti aging, skin rejuvenation, body contouring",
     "Different league on links, different buyer. Useful only as a benchmark"),
    ("Burlington Medical Aesthetics", "burlingtonmedicalaesthetics.com", "Adjacent market", "Burlington, with an Oakville landing page",
     "Physician-led, ranks into Oakville from a dedicated /oakville page",
     "Medium", "Botox oakville, med spa oakville",
     "Proof that a single well-written location page reaches across town lines. The model for /med-spa-burlington"),
    ("Bar Beauty Medical", "barbeauty.ca", "Adjacent market", "Toronto, with an Oakville landing page",
     "RN injectors, published prices, targets Oakville from downtown Toronto",
     "Watch", "Oakville medical aesthetics",
     "Transparent pricing is what earns them the click from this far away"),
]

# These are not clinics, but they hold the slots above the clinics on most
# of the head terms. Getting listed on them is a citations job, not a
# content job — the Local Citations tab is where that work lives.
DIRECTORIES = [
    ("RealSelf", "realself.com", "Directory", "Ranks for top botox providers in Oakville"),
    ("Yelp Canada", "yelp.ca", "Directory", "Owns the TOP 10 BEST botox in Oakville result"),
    ("WhatClinic", "whatclinic.com", "Directory", "Ranks for medical aesthetics clinics in Oakville"),
    ("Wheree", "wheree.com", "Directory", "Scrapes Google listings — yours will appear whether you claim it or not"),
    ("BestInRatings", "bestinratings.com", "Listicle", "Top 5 botox clinics in Oakville. Roundups like this are pitchable"),
]

# ------------------------------------------------------------------- styling

def hdr(ws, row, headers, widths):
    for i, (h, w) in enumerate(zip(headers, widths), start=1):
        c = ws.cell(row=row, column=i, value=h)
        c.font = Font(name="Calibri", size=10, bold=True, color=INK)
        c.fill = PatternFill("solid", fgColor=CREAM_DEEP)
        c.alignment = Alignment(vertical="center", wrap_text=True)
        c.border = Border(bottom=Side("thin", color=GOLD))
        ws.column_dimensions[get_column_letter(i)].width = w
    ws.row_dimensions[row].height = 30


def title_block(ws, title, subtitle):
    a = ws.cell(row=1, column=1, value=title)
    a.font = Font(name="Georgia", size=15, bold=True, color=INK)
    b = ws.cell(row=2, column=1, value=subtitle)
    b.font = Font(name="Calibri", size=10, italic=True, color=GREY)
    ws.row_dimensions[1].height = 22


def dv(ws, values, cells):
    v = DataValidation(type="list", formula1='"' + ",".join(values) + '"', allow_blank=True)
    ws.add_data_validation(v)
    v.add(cells)
    return v


def body(ws, row, values, wrap_cols=()):
    for i, val in enumerate(values, start=1):
        c = ws.cell(row=row, column=i, value=val)
        c.font = Font(name="Calibri", size=10, color=INK)
        c.alignment = Alignment(vertical="top", wrap_text=(i in wrap_cols))
        c.border = Border(bottom=Side("thin", color=CREAM))


def note(ws, coord, text):
    ws[coord].comment = Comment(text, "SEO roadmap")


# =============================================================== build it

wb = openpyxl.load_workbook(BOOK)

for name in ("Target Keywords", "Competitors", "Sheet1"):
    if name in wb.sheetnames:
        del wb[name]

# ------------------------------------------------------- Target Keywords tab
ws = wb.create_sheet("Target Keywords")
title_block(ws, "Target keywords — Oakville and Halton",
            "Cleo's forty, extended. One row per keyword, one page per cluster. "
            "Blank metric columns are the Semrush pull — see Start Here.")

HEADERS = ["Priority", "Keyword", "Cluster", "Location", "Intent", "Source",
           "Target page", "Page type", "Volume", "KD %", "CPC (CAD)",
           "Position", "Checked", "Status", "Notes"]
WIDTHS = [8, 40, 17, 14, 14, 11, 30, 18, 9, 8, 11, 10, 12, 13, 62]
hdr(ws, 4, HEADERS, WIDTHS)

note(ws, "I4", "Semrush: Keyword Overview, ca database. Monthly searches.")
note(ws, "J4", "Semrush keyword difficulty, 0 to 100. Under 30 is reachable inside six months.")
note(ws, "K4", "Semrush CPC. What an ad click costs — the honest read on commercial value.")
note(ws, "L4", "Current organic position. Blank means not ranking. Re-check monthly with the Metrics tab.")
note(ws, "F4", "Cleo's list = one of the forty priorities. Added = found while mapping keywords to pages.")

r = 5
seed_ids = {k[0] for k in SEED}
for pri, kw, cluster, loc, intent, page, ptype, notes in KEYWORDS:
    source = "Cleo's list" if pri in seed_ids and pri <= 40 else "Added"
    body(ws, r, [pri, kw, cluster, loc, intent, source, page, ptype,
                 None, None, None, None, None, "Not started", notes],
         wrap_cols=(15,))
    ws.cell(row=r, column=11).number_format = '$#,##0.00'
    ws.cell(row=r, column=13).number_format = 'yyyy-mm-dd'
    if source == "Cleo's list":
        ws.cell(row=r, column=6).font = Font(name="Calibri", size=10, bold=True, color=INK)
    r += 1
last = r - 1

dv(ws, STATUSES, f"N5:N{last}")
for value, colour in (("Ranking", GREEN), ("Live", AMBER), ("Not applicable", RED)):
    ws.conditional_formatting.add(f"A5:O{last}", FormulaRule(
        formula=[f'$N5="{value}"'], fill=PatternFill("solid", fgColor=colour), stopIfTrue=False))
# Difficulty reads greenest where it is easiest, once the numbers are in.
ws.conditional_formatting.add(f"J5:J{last}", ColorScaleRule(
    start_type="num", start_value=0, start_color="DDEBD8",
    mid_type="num", mid_value=35, mid_color="FBEFD4",
    end_type="num", end_value=70, end_color="F6DCDA"))

ws.freeze_panes = "C5"
ws.auto_filter.ref = f"A4:O{last}"

# The counts that matter, above the grid, out of the filter's way.
summary = [
    ("Keywords tracked", f"=COUNTA(B5:B{last})"),
    ("From Cleo's forty", f'=COUNTIF(F5:F{last},"Cleo\'s list")'),
    ("Added while mapping", f'=COUNTIF(F5:F{last},"Added")'),
    ("Ranking", f'=COUNTIF(N5:N{last},"Ranking")'),
    ("Total monthly volume", f"=SUM(I5:I{last})"),
]
col = 5
for label, formula in summary:
    a = ws.cell(row=2, column=col, value=label)
    a.font = Font(name="Calibri", size=9, color=GREY)
    b = ws.cell(row=3, column=col, value=formula)
    b.font = Font(name="Calibri", size=12, bold=True, color=INK)
    col += 1

# ----------------------------------------------------------- Competitors tab
ws = wb.create_sheet("Competitors")
title_block(ws, "Competitors — Oakville and the surrounding towns",
            "Verified from each clinic's own site or Google listing, August 2026. "
            "Blank metric columns are the Semrush pull — see Start Here.")

CHEADERS = ["#", "Competitor", "Website", "Tier", "Area", "What they win on",
            "Threat", "Overlaps on", "AS", "Organic KWs", "Traffic /mo",
            "Reviews", "Rating", "Benchmark", "Checked", "Read this first"]
CWIDTHS = [5, 34, 32, 18, 34, 52, 9, 34, 7, 12, 12, 10, 9, 13, 12, 62]
hdr(ws, 4, CHEADERS, CWIDTHS)

note(ws, "I4", "Semrush Authority Score, 0 to 100.")
note(ws, "J4", "Semrush organic keywords, ca database.")
note(ws, "K4", "Semrush estimated organic traffic per month.")
note(ws, "L4", "Google review count. Read it off their profile the same day you read your own.")
note(ws, "N4", "Roadmap task 1 asks Cleo to name three competitors to be measured against. "
                "Mark those three Yes — they go into Semrush Position Tracking and the Metrics tab.")

r = 5
for i, (name, domain, tier, area, wins, threat, overlap, notes) in enumerate(COMPETITORS, start=1):
    body(ws, r, [i, name, domain, tier, area, wins, threat, overlap,
                 None, None, None, None, None, "", None, notes],
         wrap_cols=(5, 6, 8, 16))
    ws.cell(row=r, column=15).number_format = 'yyyy-mm-dd'
    ws.cell(row=r, column=13).number_format = '0.0'
    r += 1
clast = r - 1

dv(ws, PRIORITIES, f"G5:G{clast}")
dv(ws, ["Yes", "No"], f"N5:N{clast}")
ws.conditional_formatting.add(f"G5:G{clast}", FormulaRule(
    formula=['$G5="High"'], fill=PatternFill("solid", fgColor=RED), stopIfTrue=False))
ws.conditional_formatting.add(f"G5:G{clast}", FormulaRule(
    formula=['$G5="Medium"'], fill=PatternFill("solid", fgColor=AMBER), stopIfTrue=False))
ws.conditional_formatting.add(f"A5:P{clast}", FormulaRule(
    formula=['$N5="Yes"'], fill=PatternFill("solid", fgColor=GREEN), stopIfTrue=False))

r = clast + 2
ws.merge_cells(start_row=r, start_column=1, end_row=r, end_column=16)
c = ws.cell(row=r, column=1, value="NOT CLINICS, BUT THEY HOLD THE SLOTS")
c.font = Font(name="Calibri", size=10, bold=True, color=INK)
for col in range(1, 17):
    ws.cell(row=r, column=col).fill = PatternFill("solid", fgColor=CREAM_DEEP)
ws.row_dimensions[r].height = 20
r += 1
sub = ws.cell(row=r, column=1,
              value="Directories and roundups sit above the clinics on most head terms. "
                    "Getting listed is citation work — see the Local Citations tab.")
sub.font = Font(name="Calibri", size=9, italic=True, color=GREY)
ws.merge_cells(start_row=r, start_column=1, end_row=r, end_column=16)
r += 1
for i, (name, domain, tier, why) in enumerate(DIRECTORIES, start=1):
    body(ws, r, [i, name, domain, tier, "", why], wrap_cols=(6,))
    r += 1

ws.freeze_panes = "C5"

# --------------------------------------------- pointer block on Start Here
sh = wb["Start Here"]
BAND = "THE WORDS AND THE RIVALS"
anchor = None
for row in range(1, sh.max_row + 2):
    if sh.cell(row=row, column=1).value == BAND:
        anchor = row
        break
if anchor is None:
    anchor = sh.max_row + 2

pointers = [
    ("Target Keywords", "Ninety-nine keywords, each mapped to the page that should win it. "
                        "Cleo's forty are marked; the rest were added while mapping."),
    ("Competitors", "Fourteen clinics and five directories. Threat column ranks them; "
                    "the Benchmark column is where the three from roadmap task 1 get marked."),
    ("Semrush pull", "Volume, difficulty, CPC, position, authority score and traffic are blank. "
                     "The Semrush account had no API units left when this was built. "
                     "Pull the ca database, then date it here."),
    ("Date pulled", ""),
]

c = sh.cell(row=anchor, column=1, value=BAND)
c.font = Font(name="Calibri", size=10, bold=True, color=INK)
for col in range(1, 12):
    sh.cell(row=anchor, column=col).fill = PatternFill("solid", fgColor=CREAM_DEEP)
sh.row_dimensions[anchor].height = 20

r = anchor + 1
for label, text in pointers:
    a = sh.cell(row=r, column=1, value=label)
    a.font = Font(name="Calibri", size=10, bold=True, color=INK)
    b = sh.cell(row=r, column=2, value=text)
    b.font = Font(name="Calibri", size=10, color=INK)
    b.alignment = Alignment(vertical="top", wrap_text=True)
    if str(sh.cell(row=r, column=2).coordinate) and f"B{r}:D{r}" not in [str(x) for x in sh.merged_cells.ranges]:
        sh.merge_cells(start_row=r, start_column=2, end_row=r, end_column=4)
    sh.row_dimensions[r].height = 30
    r += 1

sh.cell(row=r, column=1, value="Keywords tracked").font = Font(name="Calibri", size=10, bold=True, color=INK)
sh.cell(row=r, column=2, value=f"='Target Keywords'!E3").font = Font(name="Calibri", size=10, color=INK)
r += 1
sh.cell(row=r, column=1, value="Competitors tracked").font = Font(name="Calibri", size=10, bold=True, color=INK)
sh.cell(row=r, column=2, value=f"=COUNTA(Competitors!B5:B{clast})").font = Font(name="Calibri", size=10, color=INK)

# Keep the two new tabs beside the rest of the planning work.
order = wb.sheetnames
for name in ("Competitors", "Target Keywords"):
    order.remove(name)
    order.insert(order.index("Site Build") + 1, name)
wb._sheets = [wb[n] for n in order]

wb.save(BOOK)
print("saved", BOOK)
print("keywords:", last - 4, "competitors:", clast - 4, "directories:", len(DIRECTORIES))
print("sheets:", wb.sheetnames)
