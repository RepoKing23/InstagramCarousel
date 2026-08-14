#!/usr/bin/env python3
"""
Builds strategy/ig-content-strategy.xlsx and every content/<slug>/brief.md.

Run from the repo root:
    pip install openpyxl
    python3 strategy/build-content-calendar.py

Covers Aug 17 - Nov 30, 2026. August 17-31 are the original posts (kept, with
the prescription-drug wording corrected). Sept 1 - Nov 30 is a daily cadence,
one post per day, 91 posts, rebuilt from scratch.

Two rules hold across every post in here:

1. Health Canada restricts advertising a prescription drug's benefits to the
   public. Botox and Dysport are prescription drugs in Canada. Client-facing
   copy therefore says "anti-wrinkle injections" or "wrinkle-relaxing
   treatment", never a brand name paired with what it does. Dermal fillers are
   medical devices rather than drugs, so filler copy has more latitude.
2. The CNO advertising standard rules out guarantees, superiority claims
   ("best injector"), and misleading before-and-afters.

If you add a post, keep both rules. The Compliance line in each brief is there
so whoever designs the slide can check the artwork matches the copy.
"""

import datetime as dt
import os
import re
import shutil
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.worksheet.datavalidation import DataValidation
from openpyxl.utils import get_column_letter

REPO = os.path.dirname(os.path.dirname(os.path.abspath(__file__)))
CONTENT = os.path.join(REPO, "content")
XLSX = os.path.join(REPO, "strategy", "ig-content-strategy.xlsx")

INK, GOLD, GOLD_DARK = "2E2A23", "C9A85C", "B99753"
CREAM_DEEP, GREY = "E5DCCA", "6E6557"

WEEK_ONE = dt.date(2026, 8, 17)
REPO_URL = "https://github.com/RepoKing23/InstagramCarousel/tree/claude/luxury-beauty-seo-roadmap-3gad4b/content"

# Slot shape by weekday. Two carousels a week (Mon, Thu); everything else is a
# single image. That mix is what makes seven posts a week fit the hours.
SHAPE = {
    0: ("Carousel", "Education", "Instagram feed"),
    1: ("Single Image", "Education", "IG + FB + TikTok photo"),
    2: ("Single Image", "Engagement", "IG + FB + TikTok photo"),
    3: ("Carousel", "Trust & Proof", "Instagram feed"),
    4: ("Single Image", "Personality & BTS", "IG + FB + TikTok photo"),
    5: ("Single Image", "Engagement", "IG + FB + TikTok photo"),
    6: ("Single Image", "Promo & CTA", "IG + FB + TikTok photo"),
}

HASHTAGS = {
    "A": ("Set A: Anti-wrinkle education",
          "#antiwrinkleinjections #wrinklerelaxer #nurseinjector #aestheticnurse #naturalresults "
          "#preventativeaesthetics #facialaesthetics #oakvilleaesthetics #oakvillemedspa #haltonregion"),
    "B": ("Set B: Filler & facial balancing",
          "#dermalfiller #lipfiller #cheekfiller #facialbalancing #nurseinjector #naturallips "
          "#subtleenhancement #aestheticnurse #oakvilleaesthetics #oakvillemedspa"),
    "C": ("Set C: Trust & personality",
          "#nursepractitioner #npinjector #aestheticnurse #medspalife #honestbeauty #naturalresults "
          "#selfcare #beautyclinic #oakvilleaesthetics #oakvillemedspa"),
    "D": ("Set D: Seasonal & gifting",
          "#holidayglow #partyseasonprep #giftofglow #selfcaregift #aestheticnurse #nurseinjector "
          "#naturalresults #oakvilleaesthetics #oakvillemedspa #haltonregion"),
    "E": ("Set E: Skin health & boosters",
          "#skinboosters #skinhydration #dermaplaning #glowingskin #skinhealth #aestheticnurse "
          "#nurseinjector #naturalresults #oakvilleaesthetics #oakvillemedspa"),
}

COMPLIANT = "Compliant as written. No prescription brand name paired with a benefit claim."
CHECK_ART = ("Copy is compliant, but the finished artwork still carries a prescription brand "
             "name. Re-render the slide before this posts.")

# (date, slug, title, hook, caption, cta, hashtag set, visual, compliance note)
# August: the original posts, kept. Four had brand-name wording and are corrected here.
AUGUST = [
    ("2026-08-17", "2026-08-17-5-botox-lies-you-still", "5 Anti-Wrinkle Lies You Still Believe",
     "The truth, from a nurse injector.",
     "Half of what you have heard about wrinkle-relaxing treatment came from someone who saw bad work. "
     "Let's set the record straight. Which one did you believe? Drop it below and send this to the friend who needs slide 5.",
     "DM SMOOTH or link in bio", "A", "8 slides, editorial template. ALREADY DESIGNED.", CHECK_ART),
    ("2026-08-18", "2026-08-18-service-spotlight-botox", "Service Spotlight: Anti-Wrinkle Injections",
     "Smooth, not frozen.",
     "Softened lines, full expression, and nobody can tell why you look so rested. That is this treatment done properly.",
     "DM SMOOTH to book", "A", "Single image: treatment photo + hook overlay. ALREADY DESIGNED.", CHECK_ART),
    ("2026-08-20", "2026-08-20-5-red-flags-at-a", "5 Red Flags at a Med Spa. Run.",
     "Your face deserves better than a bargain.",
     "A menu with no consultation, prices too good to be true, and nobody asking about your medical history. "
     "If you see these, walk away. I want you safe, even if you never book with me. Save this checklist.",
     "Save this checklist", "C", "Design in editorial template", COMPLIANT),
    ("2026-08-21", "2026-08-21-behind-the-scenes-the-prep", "Behind the Scenes: The Prep",
     "Zero shortcuts. Ever.",
     "Sealed products, precise dosing, clean everything. This is the part you never see, and the part that matters most.",
     "Book with confidence. Link in bio", "C", "Single image: prep tray photo + hook overlay", COMPLIANT),
    ("2026-08-22", "2026-08-22-quote-great-work-is-invisible", "Quote: Great Work Is Invisible",
     "Brand quote card.",
     "Great work is invisible. Bad work is what gave this whole industry its reputation. Agree or disagree?",
     "Comment below", "C", "Quote card in editorial brand style", COMPLIANT),
    ("2026-08-24", "2026-08-24-your-first-visit-step-by", "Your First Visit, Step by Step",
     "Nervous about your first appointment? Here is exactly what happens.",
     "Most first-timers tell me they waited years because they did not know what to expect. "
     "Consider this your preview. Send it to the friend who keeps saying one day.",
     "Book your consult. Link in bio", "C", "Editorial template + treatment room photos", COMPLIANT),
    ("2026-08-25", "2026-08-25-3-signs-your-injector-is", "3 Signs Your Injector Is Rushing You",
     "Protect yourself before you book.",
     "If your consult is shorter than your coffee order, that is a problem. Save this and trust your gut.",
     "Save + share this", "C", "Single image: 3-point checklist card", COMPLIANT),
    ("2026-08-27", "2026-08-27-meet-np-cleo-the-face", "Meet NP Cleo: The Face Behind the Glow",
     "Real education. Real results. Real care.",
     "Before you trust someone with your face, you should know who they are. Here is my story, my training, "
     "and why I will always tell you the truth about what you need.",
     "Say hi in the comments", "C", "Portrait photos of Cleo", COMPLIANT),
    ("2026-08-28", "2026-08-28-welcome-to-the-treatment-room", "Welcome to the Treatment Room",
     "Where calm meets careful.",
     "Clean, quiet, and designed for you to feel safe asking anything. Come see it in person.",
     "Book a visit. Link in bio", "C", "Single image: treatment room photo + hook", COMPLIANT),
    ("2026-08-29", "2026-08-29-this-or-that-rested-vs", "This or That: Rested vs Frozen",
     "Which one are you actually asking for?",
     "When you picture your results, which one is it? Vote below.",
     "Vote in the comments", "C", "Two-panel comparison card", COMPLIANT),
    ("2026-08-31", "2026-08-31-preventative-botox-when-to-actually", "Preventative Treatment: When to Actually Start",
     "Your friend who never seems to age? This is why.",
     "Lines are easier to prevent than to erase. Here is how to know if your late 20s or 30s is the right time, "
     "and when it is honestly still too early. No pressure, just real timing.",
     "DM PREVENT for a consult", "A", "Editorial template", COMPLIANT),
]

# Sept 1 - Nov 30, one per day, in date order. Format follows the weekday.
# (title, hook, caption, cta, set, visual)
DAILY = [
    # ---------------------------------------- September: "Lines, Honestly"
    ("What Anti-Wrinkle Injections Actually Do", "Not frozen. Relaxed.",
     "They soften the muscle movement that folds your skin. That is it. You still frown, you still laugh, the crease just stops deepening.",
     "DM SMOOTH with questions", "A", "Single image: hook card, editorial"),
    ("This or That: Rested vs Overdone", "One of these is the goal.",
     "Two versions of the same face. Which one are you actually picturing when you say you want work done? Vote below.",
     "Vote in the comments", "A", "Two-panel comparison card"),
    ("Your First Appointment, Start to Finish", "Nobody tells you this part.",
     "Six slides on exactly what happens from the moment you walk in. No surprises, no pressure, and you can stop at any point.",
     "Save this for your first visit", "C", "6-slide carousel, editorial template"),
    ("The Prep You Never See", "This is the boring part that matters.",
     "Sealed product, checked expiry, clean field, and a plan we agreed on before anything is opened.",
     "Book with confidence. Link in bio", "C", "Single image: prep tray"),
    ("Quote: Great Work Is Invisible", "Brand quote card.",
     "If people can tell, it was too much. Agree or disagree?",
     "Comment below", "C", "Quote card"),
    ("Your Questions, Answered", "Ask me anything this week.",
     "Consults are the part people skip and the part that matters most. Send me your question and I will answer it honestly, even if the answer is not yet.",
     "DM your question", "C", "Single image: DM prompt card"),

    ("5 Myths About Wrinkle-Relaxing Treatment", "Half of this is wrong.",
     "Most of what you have heard came from someone who saw bad work, not from someone who does this properly. Eight slides, honestly.",
     "Save and share", "A", "8-slide carousel, editorial template"),
    ("The Three Areas People Ask About Most", "Forehead, frown, crow's feet.",
     "These three come up in almost every consult. Here is what each one actually is and why they behave differently.",
     "DM SMOOTH to ask", "A", "Single image: face diagram, three areas marked"),
    ("Myth vs Truth: It Will Freeze My Face", "The most common fear I hear.",
     "Frozen is a dosing problem, not an inevitable outcome. Done properly you keep your expression. Which side did you believe?",
     "Comment your answer", "A", "Two-panel myth vs truth card"),
    ("How I Decide How Much You Need", "Not a set number.",
     "Six slides on what I look at: your muscle strength, how you move, what you have had before, and what you actually want to look like.",
     "Save this before your consult", "C", "6-slide carousel"),
    ("Meet Your Injector", "Nurse practitioner, not a technician.",
     "The training matters because the assessment matters. Here is what mine is and why I will tell you no when no is the right answer.",
     "Say hi in the comments", "C", "Portrait photo + hook overlay"),
    ("Quote: Prevention Beats Erasing", "Brand quote card.",
     "It is always easier to stop a line than to soften one that has been there ten years.",
     "Save this one", "A", "Quote card"),
    ("Consults Are Free. Here Is What Happens.", "No commitment, no pressure.",
     "Twenty minutes, your face, your questions, and an honest answer about whether anything is worth doing at all.",
     "Book a consult. Link in bio", "C", "Single image: CTA card"),

    ("The Face Map", "What each area actually does.",
     "Eight slides walking the upper face: what moves, what creases, and what each area is responsible for.",
     "Save the map", "A", "8-slide carousel: annotated face diagram"),
    ("When to Start, Honestly", "There is no right age. There is a right time.",
     "If the line is there when your face is still, that is the signal. If it vanishes completely, you probably have time.",
     "DM PREVENT to talk timing", "A", "Single image: hook card"),
    ("Poll: What Should I Cover Next?", "You pick this week's topic.",
     "Dosing, timing, aftercare, or what to avoid before an appointment. Vote and I will make it.",
     "Vote below", "C", "Poll card, four options"),
    ("Aftercare: Your First 48 Hours", "Do these five things.",
     "Six slides on what to do and what to skip right after treatment. Most of it is simple and most people get one of these wrong.",
     "Save for after your appointment", "A", "6-slide carousel"),
    ("Inside the Treatment Room", "Calm on purpose.",
     "Quiet, clean, and set up so you can ask the awkward question without feeling silly.",
     "Come see it. Link in bio", "C", "Treatment room photo"),
    ("Quote: Subtle Is the Point", "Brand quote card.",
     "The goal was never for people to notice. It was for them to ask if you have been sleeping better.",
     "Comment if you agree", "C", "Quote card"),
    ("October Booking Opens Today", "Two weeks out, as always.",
     "October dates are live. Consults first, always, and I will tell you if the timing is wrong for what you want.",
     "DM BOOK for October", "C", "Single image: availability card"),

    ("What This Treatment Cannot Fix", "Some things need a different answer.",
     "Eight slides on what wrinkle-relaxing treatment does nothing for, because being sold the wrong thing is how people end up disappointed.",
     "Save and send to a friend", "A", "8-slide carousel"),
    ("Lines vs Volume: Different Problems", "This is the confusion I fix most.",
     "One is muscle movement. The other is loss of support. They need different treatments and they are not interchangeable.",
     "DM if you are not sure which", "B", "Single image: side-by-side diagram"),
    ("Ask Me Anything", "Drop your question below.",
     "Anything you have wondered but felt awkward asking. I would rather answer it here than have you guess.",
     "Comment your question", "C", "DM prompt card"),
    ("Why I Sometimes Say No", "The consults where I turn people away.",
     "Six slides. Sometimes the honest answer is that you do not need it, the timing is wrong, or what you are asking for will not do what you think.",
     "Save this", "C", "6-slide carousel"),
    ("What I Tell Every First-Timer", "The same three things, every time.",
     "Start smaller than you think. Come back in two weeks. And nothing here is permanent, which is a feature, not a flaw.",
     "DM if you are considering it", "C", "Single image: quote-style card"),
    ("Quote: No Is Protection", "Brand quote card.",
     "Saying no to a treatment you do not need is the whole job.",
     "Comment below", "C", "Quote card"),
    ("Your DMs, Answered", "This week's questions.",
     "Three real questions from your messages, answered properly. Keep them coming.",
     "DM your question", "C", "Single image: Q&A card"),

    ("How Long Results Really Last", "The honest range.",
     "Eight slides on what actually affects longevity: how much, where, your metabolism, and how often you have had it before.",
     "Save this", "A", "8-slide carousel"),
    ("The Three-Month Timeline", "Week by week, what to expect.",
     "Day 3, day 14, month 2, month 3. Here is what is normal at each point so you are not worrying at week one.",
     "Save the timeline", "A", "Single image: timeline graphic"),
    ("Does It Hurt? The Honest Answer", "You asked, so here it is.",
     "It is quick and it is a fine needle, but I am not going to tell you that you feel nothing. Here is what it actually feels like.",
     "Comment if you have wondered", "A", "Single image: hook card"),

    # ---------------------------------------- October: "Filler, Understood"
    ("Inside a Filler Appointment", "From consult to mirror.",
     "Six slides on what actually happens: the mapping, the numbing, the placement, and why I stop before you think I should.",
     "Save this", "B", "6-slide carousel"),
    ("The Consult That Changed Her Mind", "She came in for lips. She left with a plan.",
     "Sometimes what you are asking for is not what will get you the result you described. That conversation is the point of a consult.",
     "DM to book a consult", "B", "Single image: quote-style card"),
    ("Quote: Faces, Not Features", "Brand quote card.",
     "Treating one feature in isolation is how faces end up looking off. We look at the whole thing.",
     "Comment below", "B", "Quote card"),
    ("DM LIPS for October Dates", "A few spots left this month.",
     "Consult first. If filler is not the right answer for what you want, I will say so.",
     "DM LIPS", "B", "Availability card"),

    ("Filler Will Not Make You Look Like Someone Else", "The fear behind most hesitation.",
     "Eight slides on why people end up looking different, what actually causes it, and how it is avoided.",
     "Save and share", "B", "8-slide carousel"),
    ("Service Spotlight: Dermal Filler", "Support, not stuffing.",
     "Filler replaces structure that has softened. Placed properly it reads as rest, not as work.",
     "DM FILLER to ask", "B", "Single image: treatment photo + hook"),
    ("This or That: Subtle vs Obvious", "Which is the brief?",
     "Two results, same treatment, different approach. Which one are you asking for? Be honest.",
     "Vote below", "B", "Two-panel comparison"),
    ("Lip Filler: Natural vs Overdone", "The difference is the plan.",
     "Volume without a plan is how lips end up looking done. Here is what changes when there is one.",
     "DM LIPS to talk", "B", "ALREADY DESIGNED — image in this folder"),
    ("How I Map a Lip Before I Touch It", "Nothing happens until this is done.",
     "Proportion, symmetry, where your lip already sits, and what your face can carry. The mapping takes longer than the treatment.",
     "Book a consult. Link in bio", "B", "BTS photo: mapping in progress"),
    ("Quote: Subtle Is the New Statement", "Brand quote card.",
     "Nobody should be able to name what you had done.",
     "Comment if you agree", "B", "Quote card"),
    ("Consult First. Always.", "Even if you know what you want.",
     "Twenty minutes to check that what you are asking for will actually do what you are hoping.",
     "DM BOOK", "B", "CTA card"),

    ("Cheek Filler: The One Nobody Talks About", "It is doing more than you think.",
     "Eight slides on why cheek support changes the lower face, and why people who ask for lips sometimes need this instead.",
     "Save this", "B", "8-slide carousel"),
    ("Service Spotlight: Cheek Filler", "Structure, not volume.",
     "Support where the face has flattened. The result reads as rested rather than fuller.",
     "DM CHEEKS to ask", "B", "Single image: treatment photo"),
    ("Myth vs Truth: Filler Migrates Everywhere", "This one has a real answer.",
     "Migration is a placement and product problem, not something that just happens. Here is the actual explanation.",
     "Comment your question", "B", "Two-panel myth vs truth"),
    ("Is Filler Right for You? Five Questions", "Answer these before you book.",
     "Six slides. If you answer no to most of these, filler is probably not your answer and I will tell you that.",
     "Save this checklist", "B", "6-slide carousel"),
    ("Facial Balancing, Explained", "Why I look at your whole face.",
     "A lip that looks wrong is often a chin or cheek problem. Treating the lip harder does not fix it.",
     "DM to book a consult", "B", "Single image: annotated face diagram"),
    ("Quote: Good Work Looks Like Nothing Happened", "Brand quote card.",
     "That is not modesty. That is the entire standard.",
     "Comment below", "B", "Quote card"),
    ("How Much Do I Need? Ask Me.", "There is no standard amount.",
     "It depends on your face, your starting point, and what you actually want. Send me a message and we will talk it through.",
     "DM your question", "B", "DM prompt card"),

    ("Filler Dissolves. Here Is the Truth.", "Yes, it can be undone.",
     "Eight slides on what dissolving actually involves, when it is the right call, and why knowing it exists should make you feel safer.",
     "Save this", "B", "8-slide carousel"),
    ("The Filler Timeline", "Day 1 to month 12.",
     "Swelling, settling, the two-week check, and how long it genuinely lasts. Bookmark this before you book.",
     "Save the timeline", "B", "Single image: timeline graphic"),
    ("Poll: Lips, Cheeks or Jaw?", "What do you want to understand next?",
     "Vote and next week's carousel is yours.",
     "Vote below", "B", "Poll card"),
    ("Lip Aftercare: Your First 72 Hours", "Six things, in order.",
     "Six slides covering swelling, sleeping, what to avoid, and when to worry. Most of what people panic about at day two is normal.",
     "Save for after your appointment", "B", "6-slide carousel"),
    ("Aftercare Essentials I Send Home", "Everything you leave with.",
     "Written instructions, my number, and a two-week check-in already booked. Aftercare is not an afterthought.",
     "Book with confidence", "B", "Flat lay: aftercare kit"),
    ("Quote: Your Face Is Not a Trend", "Brand quote card.",
     "Trends change every eighteen months. Your face does not need to keep up.",
     "Comment below", "C", "Quote card"),
    ("November Books Are Opening", "Holiday season fills early.",
     "If you want to look rested for December, November is when that gets planned. Consults are open now.",
     "DM BOOK for November", "D", "Availability card"),

    ("5 Filler Myths, Corrected", "Starting with the big one.",
     "Eight slides on what filler does, what it does not, and where the internet has this wrong.",
     "Save and share", "B", "8-slide carousel"),
    ("Facial Balancing 101", "Where to actually start.",
     "Most people start with the feature that bothers them. Often the fix is somewhere else entirely.",
     "DM to talk it through", "B", "Single image: diagram"),
    ("Myth vs Truth: Filler Edition", "Four of them, quickly.",
     "The ones that come up in almost every consult. Which of these did you think was true?",
     "Comment your answer", "B", "Two-panel card"),
    ("Why We Treat Faces, Not Lines", "The whole philosophy in six slides.",
     "Chasing individual lines is how people end up looking treated. Reading the whole face is how they end up looking well.",
     "Save this", "B", "6-slide carousel"),
    ("A Day in the Treatment Room", "Behind the scenes.",
     "Setup, consults, the two-week checks, and the restock nobody sees. This is the actual job.",
     "Book a visit. Link in bio", "C", "BTS photo series"),
    ("Quote: Confidence Is Not a Filter", "Brand quote card.",
     "You should like your face on a Tuesday morning, not just in good lighting.",
     "Comment below", "C", "Quote card"),

    # ---------------------------------------- November: "Rested for the Season"
    ("Holiday Countdown: Book By These Dates", "Work backwards from the party.",
     "If you have an event in December, here are the dates that actually work. Leaving it to the week before is how people end up disappointed.",
     "DM BOOK", "D", "Single image: date countdown card"),
    ("Party Season Prep: What and When", "The full timeline.",
     "Eight slides on what to book, how far ahead, and what genuinely should not be done close to an event.",
     "Save this timeline", "D", "8-slide carousel"),
    ("The Six-Week Rule", "Before any event that matters.",
     "Six weeks gives time to settle, review at two weeks, and adjust if needed. Two weeks gives you swelling in photographs.",
     "DM to plan your timing", "D", "Single image: rule card"),
    ("This or That: Glow Now vs Glow for New Year", "When is your deadline?",
     "Both are fine. They are just different plans. Which one are you working towards?",
     "Vote below", "D", "Two-panel card"),
    ("What a First Consult Sounds Like", "The actual conversation.",
     "Six slides of the real questions I ask and why each one matters. Nothing here is a sales script.",
     "Save this", "C", "6-slide carousel"),
    ("The Clinic Is Holiday Ready", "Quietly festive.",
     "Same calm room, slightly better candle. Come in from the cold.",
     "Book a visit. Link in bio", "D", "Treatment room photo, seasonal"),
    ("Quote: Book the Appointment", "Brand quote card.",
     "The one you have been thinking about since March.",
     "Comment below", "D", "Quote card"),

    ("Gift Certificates Are Here", "For the person who never books anything.",
     "Any amount, no expiry pressure, and they still get a proper consult before anything happens.",
     "DM GIFT", "D", "Gift certificate card"),
    ("The Gift of Glow: What to Actually Give", "Not everything makes a good gift.",
     "Eight slides on which treatments work as a present, which really do not, and why a consult is always part of it.",
     "Save this", "D", "8-slide carousel"),
    ("Skin Boosters: The Quiet One", "No volume, no movement change.",
     "Hydration placed into the skin itself. It does not change your shape. It changes how your skin holds light.",
     "DM GLOW to ask", "E", "Single image: treatment photo"),
    ("Ask Me Anything: Holiday Edition", "Timing questions welcome.",
     "What to book, when, and what to skip before December. Ask away.",
     "Comment your question", "D", "DM prompt card"),
    ("Three Months of Honest Answers", "What you asked most.",
     "Six slides rounding up the questions that came up again and again since August, with the short answers.",
     "Save this", "C", "6-slide carousel"),
    ("Thank You, Three Months In", "This account started in August.",
     "Thank you for the questions, the messages, and for trusting me with your face. It genuinely means something.",
     "Say hi in the comments", "C", "Portrait photo + note"),
    ("Quote: Tag Someone Who Deserves a Treat", "Brand quote card.",
     "You know exactly who.",
     "Tag them below", "D", "Quote card"),

    ("Last Call for December", "Dates are going.",
     "If December matters to you, this is the week to sort it. Consults still come first.",
     "DM BOOK", "D", "Availability card"),
    ("Skin Boosters vs Filler: Not the Same", "People mix these up constantly.",
     "Eight slides on the difference: one adds structure, one improves skin quality. Asking for the wrong one is common.",
     "Save this", "E", "8-slide carousel"),
    ("Hydration From the Inside", "Why your cream is not enough.",
     "Topical products work on the surface. Skin boosters work in the layer underneath. Different jobs.",
     "DM GLOW", "E", "Single image: diagram"),
    ("Poll: What Is Your Winter Skin Problem?", "Pick one.",
     "Dullness, dryness, texture, or tightness. Vote and I will cover the winner.",
     "Vote below", "E", "Poll card"),
    ("How I Build a Plan, Not a Menu", "Why I do not hand you a price list.",
     "Six slides on sequencing: what to do first, what can wait a year, and what you may never need at all.",
     "Save this", "C", "6-slide carousel"),
    ("Restocking Day", "The unglamorous half of the job.",
     "Checking lots, dates and seals. Nothing goes near your face that I have not personally checked in.",
     "Book with confidence", "C", "BTS photo: stock check"),
    ("Quote: Glow Is Maintenance, Not Magic", "Brand quote card.",
     "It is a plan you keep, not a thing you buy once.",
     "Comment below", "E", "Quote card"),

    ("December Is Filling Up", "Consults still open.",
     "A handful of treatment slots left before the holidays, and consults available all month.",
     "DM BOOK", "D", "Availability card"),
    ("Winter Skin: What Actually Helps", "Beyond drinking more water.",
     "Eight slides on what changes in cold months, what genuinely helps, and what is marketing.",
     "Save this", "E", "8-slide carousel"),
    ("Dermaplaning, Explained", "The one with the immediate payoff.",
     "Surface exfoliation and fine hair removal. Makeup sits better the same day, which is why people book it before events.",
     "DM SMOOTH to book", "E", "Single image: treatment photo"),
    ("This or That: Dermaplaning vs Facial", "Different jobs, both useful.",
     "One resurfaces, one hydrates and treats. Which one suits what you are after?",
     "Vote below", "E", "Two-panel card"),
    ("What I Wish Every Client Asked Me", "Five questions worth asking.",
     "Six slides of the questions that tell you whether an injector is worth trusting. Ask me these. Ask anyone these.",
     "Save and use this", "C", "6-slide carousel"),
    ("Black Friday, My Way", "No countdown timers here.",
     "I am not discounting medical treatment to create urgency. What I will do is spend longer in consults this week.",
     "DM to book a consult", "C", "Single image: statement card"),
    ("Quote: You Do Not Need Fixing", "Brand quote card.",
     "Nothing here is about repair. It is about feeling like yourself.",
     "Comment below", "C", "Quote card"),

    ("New Year, New Plan", "Consults open for January.",
     "If this year is the year you finally ask the questions, January books are open and consults cost nothing.",
     "DM BOOK for January", "C", "Availability card"),
    ("Coming in 2027", "What is next here.",
     "Eight slides on what is being added, what stays exactly as it is, and what you have asked for most.",
     "Comment what you want to see", "C", "8-slide carousel"),
]

PRESERVE = {"2026-10-08-lip-filler-natural-vs-overdone"}
THEMES = {9: "Lines, Honestly", 10: "Filler, Understood", 11: "Rested for the Season"}


def slugify(date_str, title):
    words = re.sub(r"[^a-z0-9\s-]", "", title.lower()).split()
    return f"{date_str}-" + "-".join(words)[:38].rstrip("-")


def build_rows():
    rows = []
    for d, slug, title, hook, cap, cta, hs, vis, comp in AUGUST:
        date = dt.date.fromisoformat(d)
        fmt, pillar, plat = SHAPE[date.weekday()]
        if "8 slides" in vis or "carousel" in vis.lower():
            fmt, plat = "Carousel", "Instagram feed"
        elif "Single image" in vis or "photo" in vis.lower() or "card" in vis.lower():
            fmt, plat = "Single Image", "IG + FB + TikTok photo"
        rows.append(dict(date=date, slug=slug, title=title, hook=hook, caption=cap, cta=cta,
                         hs=hs, visual=vis, compliance=comp, fmt=fmt, pillar=pillar,
                         platforms=plat, theme="Lines, Honestly (Launch)"))

    start = dt.date(2026, 9, 1)
    assert len(DAILY) == 91, f"expected 91 daily posts, got {len(DAILY)}"
    for i, (title, hook, cap, cta, hs, vis) in enumerate(DAILY):
        date = start + dt.timedelta(days=i)
        fmt, pillar, plat = SHAPE[date.weekday()]
        if title.startswith("Quote:"):
            pillar = "Engagement"
        slug = slugify(date.isoformat(), title)
        # keep the one post that already has finished artwork
        for p in PRESERVE:
            if p.startswith(date.isoformat()):
                slug, fmt, plat = p, "Single Image", "IG + FB + TikTok photo"
        rows.append(dict(date=date, slug=slug, title=title, hook=hook, caption=cap, cta=cta,
                         hs=hs, visual=vis, compliance=COMPLIANT, fmt=fmt, pillar=pillar,
                         platforms=plat, theme=THEMES[date.month]))
    return rows


def write_briefs(rows):
    keep = {r["slug"] for r in rows}
    removed = []
    for name in sorted(os.listdir(CONTENT)):
        path = os.path.join(CONTENT, name)
        if os.path.isdir(path) and name not in keep:
            shutil.rmtree(path)
            removed.append(name)

    for r in rows:
        folder = os.path.join(CONTENT, r["slug"])
        os.makedirs(os.path.join(folder, "images"), exist_ok=True)
        week = ((r["date"] - WEEK_ONE).days // 7) + 1
        set_name, tags = HASHTAGS[r["hs"]]
        brief = f"""# {r['title']}

**Post date:** {r['date'].strftime('%A, %B %-d, %Y')} (Week {week})
**Format:** {r['fmt']}
**Platforms:** {r['platforms']}
**Pillar:** {r['pillar']}
**Theme:** {r['theme']}

## Hook
{r['hook']}

## Caption
{r['caption']}

**CTA:** {r['cta']}

## Hashtags ({set_name})
{tags}

## Visual / Asset
{r['visual']}

## Compliance
{r['compliance']}

## Design
Style: Editorial (house style)
Reference: design/botox-myths.html. Cream palette, Playfair Display headlines,
Pinyon Script accents, Jost labels, LB logo from design/brand/.
To use a different look for this post, replace this section with a description
or drop an inspiration image in this folder.

---
Drop the final image(s) for this post in the images/ subfolder.
"""
        with open(os.path.join(folder, "brief.md"), "w") as f:
            f.write(brief)
    return removed


def write_xlsx(rows):
    wb = Workbook()

    def hdr(ws, row, headers, widths):
        for i, (h, w) in enumerate(zip(headers, widths), start=1):
            c = ws.cell(row=row, column=i, value=h)
            c.font = Font(name="Calibri", size=10, bold=True, color=INK)
            c.fill = PatternFill("solid", fgColor=CREAM_DEEP)
            c.alignment = Alignment(vertical="center", wrap_text=True)
            c.border = Border(bottom=Side("thin", color=GOLD))
            ws.column_dimensions[get_column_letter(i)].width = w
        ws.row_dimensions[row].height = 26

    # ---- How to Use
    ws = wb.active
    ws.title = "How to Use"
    ws.cell(row=1, column=1, value="Luxury Beauty by Cleo R.").font = Font(name="Georgia", size=15, bold=True, color=INK)
    ws.cell(row=2, column=1, value=f"Instagram Content Strategy. Aug 17 to Nov 30, 2026 ({len(rows)} posts, one per day from Sept 1).").font = Font(name="Calibri", size=10, italic=True, color=GREY)
    lines = [
        ("", ""),
        ("THE WEEK", ""),
        ("Monday", "Education carousel, 8 slides. The anchor post of the week."),
        ("Tuesday", "Single image, education. Cross-post to Facebook and TikTok."),
        ("Wednesday", "Single image, engagement. Poll, this-or-that, or myth vs truth."),
        ("Thursday", "Carousel, 6 slides. Trust, proof or process."),
        ("Friday", "Single image, personality and behind the scenes."),
        ("Saturday", "Quote card. Lowest effort day by design."),
        ("Sunday", "Single image, promo or call to action."),
        ("", ""),
        ("WHY THIS SHAPE", ""),
        ("Two carousels a week", "Carousels take roughly 90 minutes each. Five singles take about 30. That is what makes daily fit."),
        ("Saturday is cheap", "Quote cards are a template swap. Protect that, it is what keeps the week survivable."),
        ("Batch it", "Design Monday and Thursday first each week. The rest follows quickly."),
        ("", ""),
        ("THE TWO RULES", ""),
        ("Health Canada", "Botox and Dysport are prescription drugs here. Never pair a brand name with what it does. Say anti-wrinkle injections."),
        ("CNO standard", "No guarantees, no best-injector claims, no misleading before-and-afters."),
        ("Every brief", "carries a Compliance line. Check the artwork matches it before anything posts."),
        ("", ""),
        ("HOW TO WORK THIS SHEET", ""),
        ("1. Plan", "Everything is pre-filled in Posting Schedule. Adjust topics freely, keep the two rules."),
        ("2. Design", "Ask Claude to build any post from its row."),
        ("3. Files", "Every post has a folder in the repo. Click its Folder Link for the brief and final images."),
        ("4. Track", "Update Status as each post moves: Idea, Design, Ready, Posted."),
        ("5. Measure", "After 48 hours fill Reach, Saves, Comments, DMs, Bookings from IG Insights."),
        ("6. Repeat winners", "Anything with high saves or DMs becomes next month's sequel."),
        ("", ""),
        ("YOU EDIT", "Status, Posted Link, and the performance columns. Everything else is a draft you may rewrite."),
        ("", ""),
        ("STATUS SNAPSHOT", ""),
    ]
    r = 4
    for a, b in lines:
        ca = ws.cell(row=r, column=1, value=a)
        cb = ws.cell(row=r, column=2, value=b)
        ca.font = Font(name="Calibri", size=10, bold=True, color=(GOLD_DARK if a.isupper() and a else INK))
        cb.font = Font(name="Calibri", size=10, color=INK)
        cb.alignment = Alignment(wrap_text=True, vertical="top")
        r += 1
    n = len(rows)
    for lbl, f in [("Total posts planned", f"=COUNTA('Posting Schedule'!H2:H{n+1})"),
                   ("Ideas", f'=COUNTIF(\'Posting Schedule\'!N2:N{n+1},"Idea")'),
                   ("In design", f'=COUNTIF(\'Posting Schedule\'!N2:N{n+1},"Design")'),
                   ("Ready to post", f'=COUNTIF(\'Posting Schedule\'!N2:N{n+1},"Ready")'),
                   ("Posted", f'=COUNTIF(\'Posting Schedule\'!N2:N{n+1},"Posted")')]:
        ws.cell(row=r, column=1, value=lbl).font = Font(name="Calibri", size=10, bold=True, color=INK)
        ws.cell(row=r, column=2, value=f).font = Font(name="Calibri", size=10, color=INK)
        r += 1
    ws.column_dimensions["A"].width = 24
    ws.column_dimensions["B"].width = 104

    # ---- Posting Schedule
    ws = wb.create_sheet("Posting Schedule")
    heads = ["Date", "Day", "Week", "Month Theme", "Format", "Platforms", "Pillar", "Topic & Hook",
             "Caption Draft", "CTA", "Hashtag Set", "Visual / Asset", "Folder Link", "Status",
             "Posted Link", "Reach", "Saves", "Comments", "DMs", "Bookings"]
    widths = [12, 6, 6, 20, 13, 20, 17, 46, 60, 22, 11, 40, 46, 10, 14, 8, 8, 10, 7, 10]
    hdr(ws, 1, heads, widths)
    for i, r_ in enumerate(rows, start=2):
        week = ((r_["date"] - WEEK_ONE).days // 7) + 1
        vals = [r_["date"], r_["date"].strftime("%a"), week, r_["theme"], r_["fmt"], r_["platforms"],
                r_["pillar"], f"{r_['title']} | {r_['hook']}",
                r_["caption"] + "\n\n" + HASHTAGS[r_["hs"]][1], r_["cta"], f"Set {r_['hs']}",
                r_["visual"], f"{REPO_URL}/{r_['slug']}", "Idea", None, None, None, None, None, None]
        for col, v in enumerate(vals, start=1):
            c = ws.cell(row=i, column=col, value=v)
            c.font = Font(name="Calibri", size=10, color=INK)
            c.alignment = Alignment(vertical="top", wrap_text=(col in (8, 9, 12, 13)))
            if col == 1:
                c.number_format = "yyyy-mm-dd"
    last = len(rows) + 1
    ws.freeze_panes = "A2"
    ws.auto_filter.ref = f"A1:T{last}"
    v = DataValidation(type="list", formula1='"Idea,Design,Ready,Posted"', allow_blank=True)
    ws.add_data_validation(v)
    v.add(f"N2:N{last}")

    # ---- Hashtag Sets
    ws = wb.create_sheet("Hashtag Sets")
    hdr(ws, 1, ["Set", "Use for", "Tags"], [8, 34, 96])
    for i, (k, (name, tags)) in enumerate(HASHTAGS.items(), start=2):
        for col, val in enumerate([f"Set {k}", name.split(": ", 1)[1], tags], start=1):
            c = ws.cell(row=i, column=col, value=val)
            c.font = Font(name="Calibri", size=10, color=INK)
            c.alignment = Alignment(wrap_text=True, vertical="top")
    ws.cell(row=len(HASHTAGS) + 3, column=1,
            value="No #botox or brand-name tags anywhere. Local tags do the work: #oakvilleaesthetics, #oakvillemedspa, #haltonregion.").font = Font(name="Calibri", size=10, italic=True, color=GREY)

    # ---- Compliance
    ws = wb.create_sheet("Compliance")
    ws.cell(row=1, column=1, value="What we can and cannot say").font = Font(name="Georgia", size=14, bold=True, color=INK)
    hdr(ws, 3, ["Rule", "Why", "Instead of", "Write"], [26, 40, 34, 40])
    rules = [
        ("No prescription brand name with a benefit claim",
         "Health Canada restricts advertising prescription drug benefits to the public. Botox and Dysport are prescription drugs in Canada.",
         "Botox softens lines and leaves you rested", "Anti-wrinkle injections soften the movement that creases skin"),
        ("No brand-name hashtags", "A hashtag is advertising too.", "#botox #botoxbeforeandafter", "#antiwrinkleinjections #wrinklerelaxer"),
        ("No guarantees", "CNO advertising standard.", "Guaranteed results", "What results usually look like, and what affects them"),
        ("No superiority claims", "CNO advertising standard.", "Oakville's best injector", "A nurse practitioner who will tell you no"),
        ("Honest before-and-afters", "CNO advertising standard.", "Untouched miracle result", "Consented, unedited, with the treatment and timeframe stated"),
        ("Fillers have more latitude", "Dermal fillers are medical devices, not prescription drugs, so naming them is permitted.", "-", "Brand names acceptable for fillers; still no guarantees or superiority claims"),
    ]
    for i, row in enumerate(rules, start=4):
        for col, val in enumerate(row, start=1):
            c = ws.cell(row=i, column=col, value=val)
            c.font = Font(name="Calibri", size=10, color=INK)
            c.alignment = Alignment(wrap_text=True, vertical="top")
        ws.row_dimensions[i].height = 46

    wb.save(XLSX)


if __name__ == "__main__":
    rows = build_rows()
    removed = write_briefs(rows)
    write_xlsx(rows)
    carousels = sum(1 for r in rows if r["fmt"] == "Carousel")
    print(f"{len(rows)} posts: {rows[0]['date']} to {rows[-1]['date']}")
    print(f"  {carousels} carousels, {len(rows)-carousels} single images")
    print(f"  removed {len(removed)} stale folders")
    print(f"  wrote {XLSX}")
