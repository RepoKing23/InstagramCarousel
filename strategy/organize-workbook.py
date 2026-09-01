"""Apply the house formatting to every tab of the planner.

    python3 strategy/organize-workbook.py

Idempotent, so it is safe to re-run. Run it after design/build-gbp-daily.py,
which regenerates the GBP Daily tab from scratch and therefore drops styling.

It does not touch cell content except for two known data fixes, and it leaves
every hyperlink and formula alone.
"""
import openpyxl
from openpyxl.styles import Alignment, Font, PatternFill
from openpyxl.utils import get_column_letter
import os

BOOK = os.path.join(os.path.dirname(os.path.abspath(__file__)),
                    'ig-content-strategy.xlsx')

INK, CREAM, PAPER = '2E2A23', 'F3EEE5', 'FFFFFF'
INPUT_FILL = 'FFF6DF'          # cells the user fills in
BODY = 'Calibri'               # the workbook's existing face, kept on purpose

# How to Use is prose, not a table: no header band, no filter, no freeze.
PROSE = {'How to Use'}

# tab -> (freeze cell, autofilter?, colour)
LAYOUT = {
    'How to Use':       (None,  False, INK),
    'Posting Schedule': ('C2',  True,  'C9A85C'),
    'Reels & Video Routine': ('D2', True, 'C9A85C'),
    'GBP Daily':        ('C2',  True,  'C9A85C'),
    'Monthly Themes':   ('A2',  True,  '8A8578'),
    'Content Pillars':  ('A2',  True,  '8A8578'),
    'Hashtag Sets':     ('A2',  True,  '8A8578'),
    'Photo Library':    ('A2',  True,  '8A8578'),
    'Ideas Backlog':    ('A2',  True,  '8A8578'),
}

# columns the user is meant to type into, shaded so they read as inputs
INPUTS = {
    'Posting Schedule': ['Status', 'Posted Link', 'Reach', 'Saves', 'Comments',
                         'DMs', 'Bookings'],
    'Reels & Video Routine': ['Filmed', 'Edited', 'Status', 'Posted Link',
                             'Views', 'Saves', 'Shares', 'DMs', 'Bookings'],
    'GBP Daily':        ['Status', 'Notes'],
    'Photo Library':    ['Consent', 'Client Notes'],
    'Ideas Backlog':    ['Notes'],
}


def fix_data(wb):
    """Two real errors, not formatting.

    The Aug 17 and Aug 18 rows had a raw image URL sitting in GBP CTA, where
    every other row carries the CTA wording. The URL is already correct in the
    GBP Image Link column beside it, so the stray copy is just wrong.
    """
    ws = wb['Posting Schedule']
    head = [c.value for c in ws[1]]
    cta = head.index('GBP CTA') + 1
    fixed = 0
    for r in range(2, ws.max_row + 1):
        v = ws.cell(r, cta).value
        if isinstance(v, str) and v.startswith('http'):
            ws.cell(r, cta).value = 'Learn more > Instagram'
            fixed += 1
    return fixed


def style_sheet(ws, freeze, autofilter, colour):
    ws.sheet_properties.tabColor = colour
    if ws.max_row < 2 or ws.title in PROSE:
        return
    head = [c.value for c in ws[1]]

    for c in ws[1]:
        c.font = Font(name=BODY, size=11, bold=True, color=PAPER)
        c.fill = PatternFill('solid', fgColor=INK)
        c.alignment = Alignment(vertical='center', wrap_text=True)
    ws.row_dimensions[1].height = 30

    inputs = {head.index(n) + 1 for n in INPUTS.get(ws.title, []) if n in head}
    fill = PatternFill('solid', fgColor=INPUT_FILL)
    for r in range(2, ws.max_row + 1):
        for i in inputs:
            cell = ws.cell(r, i)
            if not cell.value:                 # never shade over real content
                cell.fill = fill

    if freeze:
        ws.freeze_panes = freeze
    if autofilter:
        ws.auto_filter.ref = f'A1:{get_column_letter(ws.max_column)}{ws.max_row}'


def main():
    wb = openpyxl.load_workbook(BOOK)
    fixed = fix_data(wb)
    for name, (freeze, af, colour) in LAYOUT.items():
        if name in wb.sheetnames:
            style_sheet(wb[name], freeze, af, colour)
    wb.save(BOOK)
    print(f'styled {len(LAYOUT)} tabs, repaired {fixed} GBP CTA cells')


if __name__ == '__main__':
    main()
