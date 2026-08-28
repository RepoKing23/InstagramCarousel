# -*- coding: utf-8 -*-
"""Build the delivered workbook from the exported Google Sheet.

Two additions, both of them additive:
  1. GBP Daily gains the Aug 17 and Aug 18 posts, which went out but never got
     a row. They are inserted in date order at the top.
  2. A new 'Reels & Video Routine' tab: the six reels, carousels and images
     already posted (Aug 17 to Aug 28), then the plan through week 24, in one
     chronological log.
No pre-existing cell value is overwritten, and no status is changed.
"""
import csv, datetime, openpyxl
from copy import copy
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.worksheet.datavalidation import DataValidation
from openpyxl.worksheet.hyperlink import Hyperlink
from openpyxl.utils import get_column_letter as L

SCRATCH = "/tmp/claude-0/-home-user-InstagramCarousel/2d398c26-b4fd-5582-a06c-decfe796759f/scratchpad"
SRC = f"{SCRATCH}/master.xlsx"
CSV = "/home/user/InstagramCarousel/strategy/reels-video-weekly-routine-w2-w24.csv"
POSTED = "/home/user/InstagramCarousel/strategy/reels-video-posted-log.csv"
OUT = "/home/user/InstagramCarousel/strategy/luxury-beauty-content-plan-with-reels-routine.xlsx"

# Conventions lifted from the existing tabs.
INK, BAND, RULE, GOLD = "FF2E2A23", "FFF3EEE5", "FFD8D0C0", "FFC9A85C"
hdr_font  = Font(name="Calibri", size=11, bold=True, color="FFFFFFFF")
hdr_fill  = PatternFill("solid", fgColor=INK)
hdr_align = Alignment(horizontal="center", vertical="center", wrap_text=True)
body_font = Font(name="Arial", size=10)
body_align = Alignment(vertical="top", wrap_text=True)
band_fill = PatternFill("solid", fgColor=BAND)
edge = Side(style="thin", color=RULE)
box  = Border(left=edge, right=edge, top=edge, bottom=edge)

REPO = "https://github.com/RepoKing23/InstagramCarousel"
BRANCH = "claude/med-spa-carousel-ideas-sim9wz"
RAW = f"https://raw.githubusercontent.com/RepoKing23/InstagramCarousel/{BRANCH}"
IG = "https://www.instagram.com/luxury_beauty_aestheticz/"

wb = openpyxl.load_workbook(SRC)

# ---------------------------------------------------------------- 1. GBP rows
# Both posts are marked Done on Posting Schedule and both already have a
# gbp.jpg in their post folder, but neither ever got a GBP Daily row.
BACKFILL = [
    dict(
        date=datetime.datetime(2026, 8, 17), day="Mon",
        source="IG: 5 Botox Lies You Still Believe", pillar="Education",
        short=("Five things about Botox that are not true: it freezes your face, it is addictive, "
               "you should wait until you have wrinkles, and cheap units are the same product. "
               "The honest version, from a nurse practitioner in Oakville. Full list on Instagram."),
        long=("Five things about Botox that are not true, from a nurse practitioner who injects in Oakville every week.\n\n"
              "It freezes your face. That is dosing, not the product. Placed properly you still frown, squint and smile.\n\n"
              "It is addictive. What people get used to is looking rested, not the appointment.\n\n"
              "Wait until you have wrinkles. The earlier conversation is usually the cheaper one, because a line that "
              "has not set into the skin takes less to soften.\n\n"
              "Cheap units are the same product. The vial might be. The person holding it, the dose and the plan are not.\n\n"
              "Serving Oakville, Burlington, Milton and Mississauga. Injections are done by a nurse practitioner and never delegated."),
        folder="2026-08-17-5-botox-lies-you-still",
    ),
    dict(
        date=datetime.datetime(2026, 8, 18), day="Tue",
        source="IG: Service Spotlight, Botox", pillar="Education",
        short=("Botox, explained plainly. Small doses placed in the muscles that fold your skin, so lines soften "
               "while the expression stays yours. Softened lines, full expression, and nobody can tell why you "
               "look so rested. Consultations are free."),
        long=("Botox, explained plainly, by a nurse practitioner in Oakville.\n\n"
              "Small doses are placed into the muscles that fold your skin. The muscle relaxes, the fold softens, and "
              "the line has less to press into. Done well you keep your expression. You should still frown, squint and "
              "smile afterwards.\n\n"
              "The dose is decided after watching your face move, not read off a price list, because muscle strength "
              "differs from one person to the next.\n\n"
              "Most people see the change from day three, with the full effect around day fourteen. It lasts roughly "
              "three to four months, and the first round often fades a little sooner than the ones after it.\n\n"
              "Serving Oakville, Burlington, Milton and Mississauga. Consultations are free."),
        folder="2026-08-18-service-spotlight-botox",
    ),
]
NOTE = ("Posted. Added retroactively: this tab was started on Aug 19, so the first two days were never "
        "logged. No image link on purpose. Copy reconstructed from that day's Instagram post, so replace "
        "it if what you published to GBP differed.")

g = wb["GBP Daily"]
NCOL = 17
first, n_existing = 2, 14
assert g.cell(first, 1).value == datetime.datetime(2026, 8, 19)
assert g.cell(first + n_existing, 1).value is None

# Snapshot the existing rows, then rewrite them lower down. openpyxl's
# insert_rows does not carry hyperlinks or styles, so this is done by hand.
snap = []
for r in range(first, first + n_existing):
    row = []
    for c in range(1, NCOL + 1):
        cell = g.cell(r, c)
        row.append((cell.value, copy(cell._style), cell.hyperlink))
    snap.append(row)

shift = len(BACKFILL)
for i, row in enumerate(snap):
    for c, (val, style, hl) in enumerate(row, start=1):
        cell = g.cell(first + shift + i, c)
        cell.value, cell._style = val, style
        cell.hyperlink = (Hyperlink(ref=cell.coordinate, target=hl.target,
                                    tooltip=hl.tooltip, display=hl.display) if hl else None)

# Style template: an IG-sourced (unshaded) existing row, now at its new home.
tmpl = [copy(g.cell(first + shift + 1, c)._style) for c in range(1, NCOL + 1)]

for i, p in enumerate(BACKFILL):
    r = first + i
    # Image File, Open Folder, Open Image, Preview and Photo Credit stay empty:
    # these two went out before the tab existed and are logged, not re-published.
    vals = [
        p["date"], p["day"], p["source"], p["pillar"],
        p["short"], len(p["short"]), p["long"], len(p["long"]),
        "Learn more", "@luxury_beauty_aestheticz", None,
        None, None, None, None, "Published", NOTE,
    ]
    links = {10: IG}
    for c, v in enumerate(vals, start=1):
        cell = g.cell(r, c)
        cell.value, cell._style = v, tmpl[c - 1]
        # clear whatever link the overwritten cell carried, then set ours
        cell.hyperlink = (Hyperlink(ref=cell.coordinate, target=links[c])
                          if c in links else None)
    g.row_dimensions[r].height = 132

g.auto_filter.ref = f"A1:{L(NCOL)}{first + shift + n_existing - 1}"

# --------------------------------------------------- 2. Reels & Video Routine
WIDTHS = [7, 12, 6, 18, 20, 16, 22, 16, 40, 34, 60, 34, 55, 24, 10,
          40, 32, 38, 9, 9, 12, 18, 8, 8, 8, 8, 9]
rows = list(csv.reader(open(CSV, encoding="utf-8")))
header, planned = rows[0], rows[1:]
log = list(csv.reader(open(POSTED, encoding="utf-8")))
assert log[0] == header, "posted log must use the same columns as the plan"
posted = log[1:]
assert len(header) == len(WIDTHS) == 27 and len(planned) == 69 and len(posted) == 6

def as_date(row):
    return datetime.datetime.strptime(row[1], "%b %d, %Y")

# One chronological log: what went out, then what is still planned. On a date
# carrying both, the posted row sorts first.
data = sorted(posted + planned, key=lambda r: (as_date(r), r[20] != "Posted"))
STATUS_COL = 21

ws = wb.create_sheet("Reels & Video Routine", index=2)   # after Posting Schedule
ws.sheet_properties.tabColor = GOLD
ws.sheet_view.showGridLines = False

for i, name in enumerate(header, start=1):
    c = ws.cell(1, i, name)
    c.font, c.fill, c.alignment, c.border = hdr_font, hdr_fill, hdr_align, box
    ws.column_dimensions[L(i)].width = WIDTHS[i - 1]
ws.row_dimensions[1].height = 30

for r, row in enumerate(data, start=2):
    week = int(row[0])
    for i, val in enumerate(row, start=1):
        if i == 1:
            val = week
        elif i == 2:
            val = datetime.datetime.strptime(val, "%b %d, %Y")
        c = ws.cell(r, i, val if val != "" else None)
        c.font, c.alignment, c.border = body_font, body_align, box
        if week % 2 == 0:                       # one shade per week block
            c.fill = band_fill
        if i == 2:
            c.number_format = 'mmm\\ d", "yyyy'
        if i in (1, 3, 15, 19, 20, 21):
            c.alignment = Alignment(horizontal="center", vertical="top", wrap_text=True)
        if i == STATUS_COL and val == "Posted":
            c.font = Font(name="Arial", size=10, bold=True)
    ws.row_dimensions[r].height = 75

last = len(data) + 1
ws.freeze_panes = "C2"
ws.auto_filter.ref = f"A1:{L(len(header))}{last}"
status_dv = DataValidation(type="list", formula1='"Idea,Design,Ready,Hold,Posted"', allow_blank=True)
ws.add_data_validation(status_dv); status_dv.add(f"U2:U{last}")
done_dv = DataValidation(type="list", formula1='"Yes,No"', allow_blank=True)
ws.add_data_validation(done_dv); done_dv.add(f"S2:T{last}")

# ------------------------------------------------- 3. note on 'How to Use'
h = wb["How to Use"]
assert all(h.cell(r, c).value is None for r in range(18, 34) for c in (2, 3))

def band_row(row, text):
    for col in (2, 3):
        c = h.cell(row, col)
        c.value = text if col == 2 else None
        c.font = Font(name="Arial", size=10, bold=True, color="FFFFFFFF")
        c.fill = PatternFill("solid", fgColor=INK)

def pair(row, label, value):
    a = h.cell(row, 2, label); a.font = Font(name="Arial", size=10, bold=True)
    b = h.cell(row, 3, value); b.font = Font(name="Arial", size=10)
    b.alignment = Alignment(vertical="top", wrap_text=True)

band_row(19, "REELS & VIDEO ROUTINE")
pair(20, "The tab", "Reels & Video Routine. One chronological log: what has already gone out, then the plan through week 24, Jan 31 2027, which is month 6. Nothing on Posting Schedule was changed.")
pair(21, "Wednesday", "Slot 1. Reel. Wednesday is open on the Posting Schedule.")
pair(22, "Friday", "Slot 2. Image post or paid ad creative. Weeks 2 to 13 this is the boosted version of that week's work; weeks 14 to 24 it is a fresh image post.")
pair(23, "Sunday", "Slot 3. Reel. Sunday is open on the Posting Schedule.")
pair(24, "Cadence", "Three create-and-post actions every week. 69 items in total.")
pair(25, "New themes", "The plan stopped at Nov 15. Gift Season Opens (Nov 16-30), Glow Into the New Year (December), New Year, Real Results (January) carry it to month 6.")
pair(26, "Reuse", "About half the routine repurposes finished carousels and Photo Library pairs. The Source Assets & Notes column names the exact file or post.")
pair(27, "Consent", "Rows using client photos name the pair and its consent state. Pair D is excluded while its direction is unconfirmed. Jan 24 needs new consent and photos taken over time.")
pair(28, "Already posted", "Rows with Slot 'Posted - ...' and a bold Posted status are real history, six of them, Aug 17 to Aug 28. Filter the Status column to see just those. Captions are recorded exactly as supplied; the production columns are blank because these are finished posts, not briefs.")
band_row(30, "GBP DAILY BACKFILL")
pair(31, "Added", "Aug 17 and Aug 18, marked Published. This tab was started on Aug 19, so the first two days of the plan were never logged. It now starts on Aug 17 with everything else.")
pair(32, "No image links", "Left blank on those two rows on purpose. They are a record of posts already made, not something to publish again. Their Notes cell says the copy is reconstructed from that day's Instagram post.")
for r in list(range(20, 29)) + [31, 32]:
    h.row_dimensions[r].height = 28

wb.save(OUT)
print("wrote", OUT)
