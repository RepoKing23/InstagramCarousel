#!/usr/bin/env python3
"""
Builds strategy/seo-roadmap-tracker.xlsx — the working task tracker for the
six-month local SEO, website and Instagram plan in
strategy/seo-roadmap-6-month.docx.

Run from the repo root:
    pip install openpyxl
    python3 strategy/build-seo-tracker.py

Designed to be uploaded to Google Sheets and updated daily. Every task here
traces back to a section of the roadmap document; if you change one, change
the other.

The engagement runs 26 weeks, Mon 17 Aug 2026 to Sun 14 Feb 2027. Weeks run
Monday to Sunday and every task is due the Sunday of its week. The website
build is approved — there is no path branching in this file.
"""

import datetime as dt
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.worksheet.datavalidation import DataValidation
from openpyxl.formatting.rule import FormulaRule
from openpyxl.utils import get_column_letter

OUT = "strategy/seo-roadmap-tracker.xlsx"

# Palette carried over from the roadmap document and design/botox-myths.html
INK = "2E2A23"
GOLD = "C9A85C"
CREAM_DEEP = "E5DCCA"
CREAM = "F3EEE5"
GREEN = "DDEBD8"
AMBER = "FBEFD4"
RED = "F6DCDA"
GREY = "6E6557"

STATUSES = ["Not started", "In progress", "Blocked", "Done", "Not applicable"]
OWNERS = ["Agency", "Cleo", "Both"]
PRIORITIES = ["High", "Medium", "Low"]
CADENCES = ["One-off", "Daily", "Weekly", "Fortnightly", "Monthly", "Ongoing"]
DAY_MARKS = ["Done", "Missed", "N/A"]

START = dt.date(2026, 8, 17)          # Monday
WEEKS = 26
END = START + dt.timedelta(weeks=WEEKS) - dt.timedelta(days=1)   # Sun 14 Feb 2027

# Week N is due on the Sunday that closes it.
WEEK_DUE = {w: START + dt.timedelta(weeks=w - 1, days=6) for w in range(1, WEEKS + 1)}

# Months are four-week blocks, except months five and six which run five weeks
# so the holiday fortnight sits inside one of them rather than straddling two.
MONTH_OF_WEEK = {}
for w in range(1, WEEKS + 1):
    if w <= 4:    MONTH_OF_WEEK[w] = 1
    elif w <= 8:  MONTH_OF_WEEK[w] = 2
    elif w <= 12: MONTH_OF_WEEK[w] = 3
    elif w <= 16: MONTH_OF_WEEK[w] = 4
    elif w <= 21: MONTH_OF_WEEK[w] = 5
    else:         MONTH_OF_WEEK[w] = 6

MONTH_LABEL = {
    1: "Month 1 — Aug 17 to Sep 13 — Google profile and citations, site design",
    2: "Month 2 — Sep 14 to Oct 11 — build the site, start the review engine",
    3: "Month 3 — Oct 12 to Nov 8 — launch, indexing and local links",
    4: "Month 4 — Nov 9 to Dec 6 — depth on the pages that earn",
    5: "Month 5 — Dec 7 to Jan 10 — hold the cadence through the holidays",
    6: "Month 6 — Jan 11 to Feb 14 — compound it, measure it, plan the next six",
}

# (week, area, task, owner, cadence, priority)
TASKS = [
    # ============================================ MONTH 1 — W1 to W4
    # ---- W1 (Aug 17-23): Google profile and the citation audit. No site work.
    (1, "Google profile", "Find out who owns the Google Business Profile and which email holds it", "Cleo", "One-off", "High"),
    (1, "Google profile", "Confirm with Allure whether you can use your own signage and suite designation", "Cleo", "One-off", "High"),
    (1, "Google profile", "Claim and verify the Google Business Profile", "Agency", "One-off", "High"),
    (1, "Google profile", "Correct the GBP address to L6M 0T9 and add the suite designation", "Agency", "One-off", "High"),
    (1, "Google profile", "Set real opening hours on GBP for every day of the week", "Agency", "One-off", "High"),
    (1, "Google profile", "Choose the primary GBP category", "Agency", "One-off", "High"),
    (1, "Google profile", "Choose secondary categories that avoid colliding with Allure's listing", "Agency", "One-off", "High"),
    (1, "Google profile", "Replace the dead website link on GBP with the booking link until the site is live", "Agency", "One-off", "High"),
    (1, "Google profile", "Upload real photos to GBP: treatment room, you, the space", "Both", "One-off", "High"),
    (1, "Citations", "Decide on one email address and one phone number to use everywhere", "Cleo", "One-off", "High"),
    (1, "Citations", "Audit every listing that mentions the practice and record the wrong NAP on the Citations tab", "Agency", "One-off", "High"),
    (1, "Citations", "Fill in THE RECORD on the Local Citations tab: one name, one address, one phone, one email", "Both", "One-off", "High"),
    (1, "Daily routine", "Start the daily routine: GBP post, Instagram, on-page, review ask, citation", "Cleo", "Daily", "High"),
    (1, "Daily routine", "Agree who does which daily slot and put it in the calendar", "Both", "One-off", "High"),
    (1, "Measurement", "Record the baseline: 12 reviews, 5.0 average, GBP calls, views and direction requests", "Agency", "One-off", "High"),
    (1, "Measurement", "Send us your real opening hours for every day of the week", "Cleo", "One-off", "High"),
    (1, "Measurement", "Name the three competitors you want to be measured against", "Cleo", "One-off", "Medium"),

    # ---- W2 (Aug 24-30): citations continue, website planning and design.
    (2, "Citations", "Correct postal code to L6M 0T9 on Birdeye", "Agency", "One-off", "High"),
    (2, "Citations", "Correct postal code to L6M 0T9 on Facebook", "Agency", "One-off", "High"),
    (2, "Citations", "Correct postal code to L6M 0T9 on Fresha", "Agency", "One-off", "High"),
    (2, "Citations", "Correct postal code to L6M 0T9 on Yelp.ca", "Agency", "One-off", "Medium"),
    (2, "Citations", "Correct postal code to L6M 0T9 on Yellow Pages", "Agency", "One-off", "Medium"),
    (2, "Citations", "Check which directories on the Local Citations build list are still live and still free", "Agency", "One-off", "Medium"),
    (2, "Website", "PLANNING WEEK: agree the sitemap and the full page inventory", "Both", "One-off", "High"),
    (2, "Website", "Build the keyword-to-page map so every page has one job", "Agency", "One-off", "High"),
    (2, "Website", "Wireframe the home page, the contact page and the treatment page template", "Agency", "One-off", "High"),
    (2, "Website", "Pull the brand kit — colours, type, logo — from design/brand", "Agency", "One-off", "Medium"),
    (2, "Website", "Send the domain registrar login for luxurybeautybycleo.ca", "Cleo", "One-off", "High"),
    (2, "Website", "Send the treatment list with current pricing", "Cleo", "One-off", "High"),
    (2, "Website", "DESIGN REVIEW with Cleo. Sign-off due Friday 28 August", "Both", "One-off", "High"),
    (2, "Instagram", "Rewrite the Instagram name field to read 'Cleo | Nurse Injector Oakville'", "Agency", "One-off", "High"),
    (2, "Instagram", "Rewrite the Instagram bio to name the town and the treatments", "Agency", "One-off", "High"),
    (2, "Instagram", "Fix the link in bio so it no longer points at the dead domain", "Agency", "One-off", "High"),
    (2, "Instagram", "Start location-tagging Oakville on every post", "Cleo", "Ongoing", "High"),
    (2, "Reviews", "Write the WhatsApp review-ask message", "Agency", "One-off", "High"),
    (2, "Reviews", "Reply to all 12 existing reviews", "Cleo", "One-off", "Medium"),

    # ---- W3 (Aug 31 - Sep 6): sign-off Monday, build starts Tuesday 1 September.
    (3, "Website", "DESIGN SIGN-OFF DUE Monday 31 August", "Cleo", "One-off", "High"),
    (3, "Website", "BUILD STARTS 1 SEPTEMBER: hosting sorted, DNS pointed, platform set up", "Agency", "One-off", "High"),
    (3, "Website", "Scaffold the page structure agreed in planning", "Agency", "One-off", "High"),
    (3, "Website", "Write the home page copy and get it approved", "Both", "One-off", "High"),
    (3, "Website", "Write the contact page copy with the corrected NAP and hours", "Agency", "One-off", "High"),
    (3, "Citations", "Correct postal code to L6M 0T9 on Apple Maps", "Agency", "One-off", "Medium"),
    (3, "Citations", "Correct postal code to L6M 0T9 on Bing Places", "Agency", "One-off", "Medium"),
    (3, "Citations", "Correct postal code to L6M 0T9 on RateMDs", "Agency", "One-off", "Low"),
    (3, "Citations", "Build the Oakville and Halton local citations from the build list", "Agency", "One-off", "High"),
    (3, "Google profile", "Fill out the GBP services list with a plain description of each treatment", "Agency", "One-off", "Medium"),
    (3, "Google profile", "Add a working booking link to GBP", "Agency", "One-off", "High"),
    (3, "Reviews", "Start sending the review ask within 24 hours of every appointment", "Cleo", "Daily", "High"),
    (3, "Compliance", "Confirm the signed before-and-after consent form is on file", "Cleo", "One-off", "High"),

    # ---- W4 (Sep 7-13)
    (4, "Website", "Home and contact pages complete in staging: NAP, hours, booking link, schema", "Agency", "One-off", "High"),
    (4, "Website", "LocalBusiness and Service schema on every page built so far", "Agency", "One-off", "High"),
    (4, "Website", "Title tags and meta descriptions for the home and contact pages", "Agency", "One-off", "High"),
    (4, "Citations", "Correct postal code to L6M 0T9 on the Canadian chamber directories", "Agency", "One-off", "Low"),
    (4, "Citations", "Claim any listing about the practice that is not under your control", "Agency", "One-off", "Medium"),
    (4, "Citations", "Verify every corrected listing now shows identical name, address and phone", "Agency", "One-off", "High"),
    (4, "Citations", "Build the healthcare and aesthetics citations, the NPAO listing included", "Agency", "One-off", "Medium"),
    (4, "Citations", "Check the College of Nurses public register entry is correct as it stands", "Agency", "One-off", "Medium"),
    (4, "Google profile", "Seed the GBP questions section with real questions and answers", "Both", "One-off", "Medium"),
    (4, "Google profile", "Post to GBP every weekday from the content calendar", "Cleo", "Daily", "High"),
    (4, "Compliance", "Compliance pass on the 21 briefs naming Botox or Dysport in client-facing copy", "Agency", "One-off", "High"),
    (4, "Compliance", "Compliance rules apply to GBP posts too — brief Cleo on what cannot be named", "Agency", "One-off", "High"),
    (4, "Measurement", "Set up the analytics property ready for launch", "Agency", "One-off", "Medium"),
    (4, "Ads", "Send us your current monthly Instagram and Google ad spend", "Cleo", "One-off", "High"),

    # ============================================ MONTH 2 — W5 to W8
    # ---- W5 (Sep 14-20)
    (5, "Website", "Build the treatment pages: lip filler, dermal filler and facial balancing", "Agency", "One-off", "High"),
    (5, "Website", "Write the copy for those pages and confirm every claim is one you can stand behind", "Both", "One-off", "High"),
    (5, "Website", "Title tag, meta description and H1 on every treatment page built", "Agency", "One-off", "High"),
    (5, "Citations", "Build the Canadian directory, maps and social profile citations", "Agency", "One-off", "Low"),
    (5, "Reviews", "Design and print the QR review card for the treatment room", "Agency", "One-off", "Medium"),
    (5, "Reviews", "Reply to every new review within two days", "Cleo", "Ongoing", "Medium"),
    (5, "Instagram", "Switch to locally weighted hashtag sets, clear of prescription brand names", "Agency", "One-off", "Medium"),
    (5, "Instagram", "Write alt text on every new post", "Cleo", "Ongoing", "Medium"),
    (5, "Content", "Publish one reel every Wednesday", "Cleo", "Weekly", "High"),

    # ---- W6 (Sep 21-27)
    (6, "Website", "Build the treatment pages: anti-wrinkle injections and skin boosters", "Agency", "One-off", "High"),
    (6, "Website", "Internal linking pass across every page built so far", "Agency", "One-off", "Medium"),
    (6, "Website", "Compress every image and write alt text across the site", "Agency", "One-off", "Medium"),
    (6, "Instagram", "Reorganise Highlights by treatment", "Agency", "One-off", "Medium"),
    (6, "Ads", "Build the local radius audience, 15km around Oakville", "Agency", "One-off", "High"),
    (6, "Ads", "Build the 30-day engagement retargeting audience", "Agency", "One-off", "High"),
    (6, "Reviews", "Check the ask against the review target and adjust it if it is lagging", "Agency", "Monthly", "High"),

    # ---- W7 (Sep 28 - Oct 4)
    (7, "Website", "Build the treatment pages: dermaplaning, electrolysis and Belkyra", "Agency", "One-off", "High"),
    (7, "Website", "Build the Oakville and surrounding-area page", "Agency", "One-off", "Medium"),
    (7, "Website", "Mobile speed pass", "Agency", "One-off", "Medium"),
    (7, "Ads", "Export the patient list from the booking system for the lookalike audience", "Cleo", "One-off", "Medium"),
    (7, "Ads", "Build the lookalike audience", "Agency", "One-off", "Medium"),
    (7, "Ads", "Launch three or four ad creatives pulled from existing carousels", "Agency", "One-off", "High"),

    # ---- W8 (Oct 5-11): launch week.
    (8, "Website", "Final content and compliance review of every page before launch", "Both", "One-off", "High"),
    (8, "Website", "LAUNCH: site live, Search Console verified, sitemap submitted, analytics running", "Agency", "One-off", "High"),
    (8, "Website", "Repoint the GBP website field at the live site", "Agency", "One-off", "High"),
    (8, "Citations", "Update every directory listing to the new website URL", "Agency", "One-off", "Medium"),
    (8, "Citations", "Update every citation built before launch from the booking link to the live site URL", "Agency", "One-off", "High"),
    (8, "Instagram", "Update the link in bio to the live site", "Agency", "One-off", "High"),
    (8, "Measurement", "Confirm Search Console and analytics are both recording on the live domain", "Agency", "One-off", "High"),
    (8, "Ads", "Fortnightly ad review: kill anything below average cost per DM, reallocate", "Agency", "Fortnightly", "High"),

    # ============================================ MONTH 3 — W9 to W12
    # ---- W9 (Oct 12-18)
    (9, "Website", "Confirm every page is indexed in Search Console and fix what is not", "Agency", "One-off", "High"),
    (9, "Website", "On-page pass on the home page: title, H1, opening copy, internal links", "Agency", "One-off", "High"),
    (9, "Local links", "Submit to the Oakville business directories", "Agency", "One-off", "Medium"),
    (9, "Local links", "Ask Allure about a cross-link between the two businesses", "Cleo", "One-off", "Low"),
    (9, "Measurement", "First post-launch read on Search Console impressions and queries", "Agency", "One-off", "Medium"),

    # ---- W10 (Oct 19-25)
    (10, "Website", "On-page pass on the four highest-intent treatment pages", "Agency", "One-off", "High"),
    (10, "Website", "Add FAQ blocks with FAQ schema to every treatment page", "Agency", "One-off", "Medium"),
    (10, "Content", "Write the first question-intent article and link it to its treatment page", "Agency", "One-off", "Medium"),
    (10, "Local links", "Approach Halton-area partners for mentions and links", "Agency", "One-off", "Medium"),
    (10, "Google profile", "Add the weight-loss service entry to GBP ahead of the launch", "Agency", "One-off", "Medium"),

    # ---- W11 (Oct 26 - Nov 1)
    (11, "Website", "Fix whatever Search Console flags: coverage, mobile usability, core web vitals", "Agency", "One-off", "High"),
    (11, "Content", "Write the second question-intent article", "Agency", "One-off", "Medium"),
    (11, "Ads", "Fortnightly ad review: kill anything below average cost per DM, reallocate", "Agency", "Fortnightly", "High"),
    (11, "Reviews", "Check the ask conversion rate and rewrite the message if it is under 30 per cent", "Agency", "One-off", "High"),

    # ---- W12 (Nov 2-8)
    (12, "Measurement", "Month three review: every number against the week-one baseline", "Agency", "Monthly", "High"),
    (12, "Measurement", "Rank check on the map pack terms and the page terms", "Agency", "Monthly", "High"),
    (12, "Content", "Write the third question-intent article", "Agency", "One-off", "Medium"),
    (12, "Local links", "Chase every unanswered link approach", "Agency", "One-off", "Low"),

    # ============================================ MONTH 4 — W13 to W16
    # ---- W13 (Nov 9-15)
    (13, "Website", "Build the weight-loss pre-launch page", "Agency", "One-off", "Medium"),
    (13, "Content", "Write the fourth article from the questions you get asked in consults", "Agency", "One-off", "Medium"),
    (13, "Instagram", "Review which posts drove profile visits and rebalance the calendar", "Agency", "Monthly", "Medium"),
    (13, "Google profile", "Refresh the GBP photo set with new room and result photos", "Both", "One-off", "Medium"),

    # ---- W14 (Nov 16-22)
    (14, "Website", "Expand the thin treatment pages: pricing ranges, aftercare, who it is not for", "Agency", "One-off", "High"),
    (14, "Local links", "Sponsor or join one Oakville community listing or event", "Cleo", "One-off", "Medium"),
    (14, "Ads", "Fortnightly ad review: kill anything below average cost per DM, reallocate", "Agency", "Fortnightly", "High"),
    (14, "Content", "Write the fifth article", "Agency", "One-off", "Medium"),

    # ---- W15 (Nov 23-29)
    (15, "Website", "Add a before-and-after gallery page using consent-cleared photos only", "Agency", "One-off", "Medium"),
    (15, "Compliance", "Compliance pass on the gallery page before it goes live", "Agency", "One-off", "High"),
    (15, "Google profile", "Post the December holiday hours to GBP", "Agency", "One-off", "Medium"),
    (15, "Reviews", "Holiday review push: ask every patient seen in November", "Cleo", "One-off", "High"),

    # ---- W16 (Nov 30 - Dec 6)
    (16, "Measurement", "Month four review: compare Search Console queries month on month", "Agency", "Monthly", "High"),
    (16, "Website", "Internal linking pass now that the articles exist", "Agency", "One-off", "Medium"),
    (16, "Content", "Write the sixth article", "Agency", "One-off", "Medium"),

    # ============================================ MONTH 5 — W17 to W21
    # ---- W17 (Dec 7-13)
    (17, "Google profile", "Set the holiday opening hours on GBP for December and January", "Agency", "One-off", "High"),
    (17, "Instagram", "Plan the December gift-card and holiday content", "Agency", "One-off", "Medium"),
    (17, "Website", "Add a gift-card or holiday offer block to the home page", "Agency", "One-off", "Medium"),

    # ---- W18 (Dec 14-20)
    (18, "Ads", "Holiday campaign: reallocate budget to the gift-card creative", "Agency", "Fortnightly", "High"),
    (18, "Content", "Write the seventh article", "Agency", "One-off", "Medium"),
    (18, "Citations", "Quarterly citation audit across both the Citations and Local Citations tabs", "Agency", "One-off", "Medium"),

    # ---- W19 (Dec 21-27)
    (19, "Instagram", "Reduced holiday cadence: stories only, one reel, but keep GBP posting daily", "Cleo", "One-off", "Medium"),
    (19, "Google profile", "Confirm the holiday hours showing on Google are actually correct", "Agency", "One-off", "High"),

    # ---- W20 (Dec 28 - Jan 3)
    (20, "Measurement", "Year-end read: reviews, map pack positions, sessions and bookings", "Agency", "Monthly", "High"),
    (20, "Content", "Plan the January content calendar", "Agency", "One-off", "Medium"),

    # ---- W21 (Jan 4-10)
    (21, "Website", "January on-page sprint: refresh every title tag and meta description", "Agency", "One-off", "High"),
    (21, "Instagram", "Back to full cadence with new-year treatment content", "Cleo", "One-off", "High"),
    (21, "Reviews", "New-year review push", "Cleo", "One-off", "High"),
    (21, "Ads", "Relaunch the standard campaign set after the holiday creative", "Agency", "Fortnightly", "High"),

    # ============================================ MONTH 6 — W22 to W26
    # ---- W22 (Jan 11-17)
    (22, "Website", "Build the full weight-loss treatment page now the service is live", "Agency", "One-off", "High"),
    (22, "Content", "Write the eighth article", "Agency", "One-off", "Medium"),
    (22, "Local links", "Second round of local link approaches", "Agency", "One-off", "Medium"),

    # ---- W23 (Jan 18-24)
    (23, "Measurement", "Full query review: find every term you rank between 5 and 15 for", "Agency", "One-off", "High"),
    (23, "Website", "Optimise the pages behind those near-miss terms", "Agency", "One-off", "High"),
    (23, "Website", "Content refresh on the three lowest-performing pages", "Agency", "One-off", "Medium"),

    # ---- W24 (Jan 25-31)
    (24, "Ads", "Fortnightly ad review: kill anything below average cost per DM, reallocate", "Agency", "Fortnightly", "High"),
    (24, "Reviews", "Check the review count against the six-month target", "Agency", "Monthly", "High"),
    (24, "Google profile", "Refresh the GBP services list and the questions section", "Agency", "One-off", "Medium"),

    # ---- W25 (Feb 1-7)
    (25, "Measurement", "Full review of every number against the week-one baseline", "Agency", "One-off", "High"),
    (25, "Website", "Technical audit: speed, schema, broken links, redirects", "Agency", "One-off", "High"),
    (25, "Citations", "Final citation audit for the engagement", "Agency", "One-off", "Medium"),

    # ---- W26 (Feb 8-14)
    (26, "Measurement", "Write the six-month report", "Agency", "One-off", "High"),
    (26, "Measurement", "Write the recommendation for the next six months", "Agency", "One-off", "High"),
    (26, "Website", "Hand over the site documentation and every login", "Agency", "One-off", "High"),
]

# ------------------------------------------------------- the daily routine
# Slot, task, cadence, what it involves, owner, where it is logged.
DAILY = [
    ("1", "Google Business Profile post",
     "Every weekday",
     "One post a day. Rotate the five types across the week: treatment spotlight, "
     "FAQ or myth, before and after, offer or availability, practice update. Copy and "
     "images come from the content calendar. Every post ends with the booking link.",
     "Cleo", "GBP Posts tab"),
    ("2", "Instagram post or stories",
     "Every weekday",
     "A feed post on the calendar days, two or three stories on the rest. Location-tag "
     "Oakville every time, write the alt text every time, use the local hashtag set.",
     "Cleo", "Content calendar"),
    ("3", "Reel",
     "Once a week — Wednesday",
     "One reel a week, filmed and cut in the Wednesday batch block. Rotate the four types "
     "across the month: a treatment in fifteen seconds, a myth-buster, behind the scenes in "
     "the room, a patient question answered to camera. Reuse each cut as a story and a GBP video.",
     "Cleo", "Daily Log tab"),
    ("4", "On-page optimisation",
     "Every weekday",
     "Before launch: keyword-to-page mapping, drafting title tags and meta descriptions, "
     "writing page copy for the build. After launch: one page improved a day on rotation — "
     "title tag, H1, internal links, image alt text, schema, thin content topped up.",
     "Agency", "Tasks tab"),
    ("5", "Review ask",
     "Every weekday",
     "The WhatsApp message inside 24 hours to every patient seen that day, and a reply to "
     "any review that landed. This is the single highest-return thing in the plan.",
     "Cleo", "Reviews tab"),
    ("6", "Citations",
     "Daily to week 4, then weekly",
     "One directory a day. The Citations tab first — the listings that exist and are wrong — "
     "then the Local Citations build list, where there is no listing at all. Once both are "
     "clean, one a week to keep them that way. Name, address and phone identical everywhere, "
     "no exceptions.",
     "Agency", "Citations + Local Citations"),
]

WEEKLY = [
    ("Mon", "Google profile insights and rank check",
     "Weekly",
     "Calls, views and direction requests for the week. Where the map pack has you for the "
     "three target terms. Two minutes, written down, every Monday.",
     "Agency", "Metrics tab"),
    ("Wed", "Content and reel batch",
     "Weekly",
     "Film the reel, batch the week's GBP posts and Instagram captions so the daily slots "
     "are a five-minute publish rather than a blank page.",
     "Both", "Content calendar"),
    ("Fri", "Log the week",
     "Weekly",
     "Fill in the week's numbers, close out the Tasks tab, flag anything blocked. "
     "Ad review every second Friday.",
     "Agency", "Tasks tab"),
]

MONTHLY = [
    ("", "Fill the Metrics tab",
     "Monthly",
     "Every number for the month end. Compare against the baseline column, not against last month.",
     "Agency", "Metrics tab"),
    ("", "Refresh the Google profile",
     "Monthly",
     "Services list, questions and answers, photo set. A profile that never changes stops earning.",
     "Agency", "Tasks tab"),
    ("", "Competitor and query check",
     "Monthly",
     "What the three named competitors are doing, and which queries Search Console shows you "
     "sitting just outside the first page for.",
     "Agency", "Metrics tab"),
    ("", "Review the month's best work",
     "Monthly",
     "The posts that drove profile visits and the pages that drove sessions. More of those next month.",
     "Both", "Metrics tab"),
]

# ------------------------------------------------- the local citation build
# THE RECORD: the canonical details pasted into every directory form. Filled in
# where the answer is already known, deliberately blank where it is not. Every
# blank here has a matching row on the Waiting On Cleo tab.
RECORD = [
    ("Business name (exact)", "Luxury Beauty by Cleo R.", "Use this spelling everywhere. No 'Med Spa', no 'Inc', no variations"),
    ("Practitioner", "Cleo Rukovo, NP", "Nurse practitioner. Some healthcare directories list the person, not the practice"),
    ("Address line 1", "3060 Preserve Drive", "Confirmed"),
    ("Suite / unit", "", "Waiting on the Allure conversation. This is what separates you from their listing"),
    ("City", "Oakville", "Confirmed"),
    ("Province", "Ontario", "Write 'Ontario' in full, or 'ON' — pick one and never mix them"),
    ("Postal code", "L6M 0T9", "NOT L6M 4L9. That is the error on most listings today"),
    ("Country", "Canada", "Confirmed"),
    ("Phone", "", "One number, everywhere. A local 905 or 289 line beats the 416 mobile if you can get one"),
    ("Email", "", "One address. Currently split between crukovo@ and luxury.beautyaestheticz@"),
    ("Website", "https://luxurybeautybycleo.ca", "Use the booking link until the site is live in October, then switch every listing"),
    ("Booking URL", "", "Needed before any directory that offers a booking button"),
    ("Primary category", "", "Set on the Google profile first, then match it everywhere else"),
    ("Secondary categories", "", "Chosen to avoid colliding with Allure's listing"),
    ("Short description (under 250 characters)", "", "Names the town and the treatments. No prescription drug names, no benefit claims"),
    ("Long description", "", "Same rules. Write once, reuse everywhere"),
    ("Hours — Monday", "", "All seven days. 'Closed' is an answer; blank is not"),
    ("Hours — Tuesday", "", ""),
    ("Hours — Wednesday", "", ""),
    ("Hours — Thursday", "", ""),
    ("Hours — Friday", "", ""),
    ("Hours — Saturday", "", ""),
    ("Hours — Sunday", "", "The only hours on record anywhere today are Sunday 12 to 5"),
    ("Year established", "", ""),
    ("Service area", "Oakville, Burlington, Milton, Mississauga, Glen Abbey, Joshua Creek, Bronte",
     "Used where a directory asks for a service radius"),
    ("Logo file", "", "From design/brand"),
    ("Cover photo", "", "The treatment room or the space, not a stock image"),
]

CITATION_TYPES = ["Local", "Healthcare", "Canadian", "Maps", "Social"]

# (directory, type, priority, where to start)
# Domains are a starting point, not gospel — Canadian directories churn, and the
# first pass on every row is confirming it still exists and still takes a free
# listing. Blank where we would rather you search than trust a stale URL.
LOCAL_CITATIONS = [
    # -- Oakville and Halton. The ones that actually say "of this town".
    ("Town of Oakville business directory", "Local", "High", "oakville.ca"),
    ("Oakville.com business listing", "Local", "High", "oakville.com"),
    ("Oakville News business listing", "Local", "Medium", "oakvillenews.org"),
    ("Visit Oakville / tourism listing", "Local", "Medium", ""),
    ("Halton Region business directory", "Local", "High", "halton.ca"),
    ("Burlington Chamber of Commerce", "Local", "Low", ""),
    ("Nextdoor business page, Oakville", "Local", "Medium", "business.nextdoor.com"),
    # -- Healthcare and aesthetics. Fewer listings, but far better matched traffic.
    ("Nurse Practitioners' Association of Ontario", "Healthcare", "High", "npao.org"),
    ("College of Nurses of Ontario public register", "Healthcare", "High", "cno.org"),
    ("Opencare", "Healthcare", "Medium", "opencare.com"),
    ("Doctify Canada", "Healthcare", "Medium", "doctify.com"),
    ("Vitals / Healthgrades Canada", "Healthcare", "Low", ""),
    ("Booksy", "Healthcare", "Medium", "booksy.com"),
    ("Vagaro", "Healthcare", "Medium", "vagaro.com"),
    ("Spas of America", "Healthcare", "Low", "spasofamerica.com"),
    # -- Canadian general directories. Low individually, useful as a consistent set.
    ("Canpages", "Canadian", "Medium", "canpages.ca"),
    ("ProfileCanada", "Canadian", "Medium", "profilecanada.com"),
    ("n49", "Canadian", "Medium", "n49.com"),
    ("Hotfrog Canada", "Canadian", "Low", "hotfrog.ca"),
    ("Canadian Business Directory", "Canadian", "Low", ""),
    ("Where To Canada", "Canadian", "Low", ""),
    ("Yably Canada", "Canadian", "Low", "yably.ca"),
    ("Data Axle Canada", "Canadian", "Medium", ""),
    # -- Maps and voice.
    ("Waze business listing", "Maps", "Medium", "waze.com"),
    ("Here WeGo", "Maps", "Low", "here.com"),
    # -- Social profiles that function as citations.
    ("LinkedIn company page", "Social", "Medium", "linkedin.com"),
    ("Pinterest business profile", "Social", "Low", "pinterest.com"),
    ("TikTok business profile", "Social", "Medium", "tiktok.com"),
]

# ------------------------------------------------------- the 90-day overview
# Ninety days from 17 August lands on 15 November, which is the end of week 13.
# So the quarter view uses the same week numbering as everything else.
CHECKPOINTS = [
    ("Day 30", dt.date(2026, 9, 15),
     "Google profile claimed, verified, complete and posting daily. All 17 existing listings "
     "corrected AND verified at L6M 0T9. Design signed off, build two weeks in. Review asks "
     "going out every day."),
    ("Day 60", dt.date(2026, 10, 15),
     "Site live. Every treatment page published, Search Console verified, sitemap submitted. "
     "Every citation repointed at the live URL. Ads running on three audiences."),
    ("Day 90", dt.date(2026, 11, 15),
     "Eight to ten new reviews. Visible in the map pack for 'nurse injector oakville'. First "
     "two or three articles live and indexed. Local link building started."),
]

# (week, phase, the one thing that matters)
QUARTER_WEEKS = [
    (1, "Foundation", "Claim and verify the Google profile. Nothing else can start until it is settled"),
    (2, "Foundation", "Design sign-off Friday 28 August, and the citation corrections begin"),
    (3, "Build", "Design signed off Monday. BUILD STARTS 1 SEPTEMBER"),
    (4, "Build", "Home and contact complete in staging. Citations verified, not just corrected"),
    (5, "Build", "First four treatment pages. The review ask is now a daily habit"),
    (6, "Build", "Second pair of treatment pages. Ad audiences built"),
    (7, "Build", "Last treatment pages and the area page. Mobile speed pass"),
    (8, "Launch", "LAUNCH WEEK. Search Console verified, sitemap in, every listing repointed"),
    (9, "Index and optimise", "Confirm every page is indexed. First local link submissions"),
    (10, "Index and optimise", "On-page pass on the four highest-intent pages. First article"),
    (11, "Index and optimise", "Fix whatever Search Console flags. Second article"),
    (12, "Index and optimise", "Month three review and a rank check against the baseline"),
    (13, "Review", "DAY 90. Full read against the week-one baseline"),
]

HARD_DATES = [
    (dt.date(2026, 8, 31), "Design sign-off", "The build gate. Everything after it moves one-for-one"),
    (dt.date(2026, 9, 1), "Build starts", "Hosting, DNS, platform, page scaffolding"),
    (dt.date(2026, 10, 5), "Launch week begins", "Site live by Sunday 11 October"),
    (dt.date(2026, 11, 15), "Day 90 review", "Every number against the week-one baseline"),
]

# ------------------------------------------------------------------- styling

def hdr(ws, row, headers, widths):
    for i, (h, w) in enumerate(zip(headers, widths), start=1):
        c = ws.cell(row=row, column=i, value=h)
        c.font = Font(name="Calibri", size=10, bold=True, color=INK)
        c.fill = PatternFill("solid", fgColor=CREAM_DEEP)
        c.alignment = Alignment(vertical="center", wrap_text=True)
        c.border = Border(bottom=Side("thin", color=GOLD))
        ws.column_dimensions[get_column_letter(i)].width = w
    ws.row_dimensions[row].height = 28


def title_block(ws, title, subtitle):
    a = ws.cell(row=1, column=1, value=title)
    a.font = Font(name="Georgia", size=15, bold=True, color=INK)
    b = ws.cell(row=2, column=1, value=subtitle)
    b.font = Font(name="Calibri", size=10, italic=True, color=GREY)
    ws.row_dimensions[1].height = 22


def dv(ws, values, cells):
    v = DataValidation(type="list", formula1='"' + ",".join(values) + '"', allow_blank=True)
    ws.add_data_validation(v)
    v.add(cells)
    return v


def band(ws, row, ncols, text):
    """A gold divider row that breaks a long grid into readable blocks."""
    ws.merge_cells(start_row=row, start_column=1, end_row=row, end_column=ncols)
    c = ws.cell(row=row, column=1, value=text)
    c.font = Font(name="Calibri", size=10, bold=True, color=INK)
    c.alignment = Alignment(vertical="center")
    for col in range(1, ncols + 1):
        ws.cell(row=row, column=col).fill = PatternFill("solid", fgColor=CREAM_DEEP)
    ws.row_dimensions[row].height = 20


def status_colours(ws, rng, col, row, done="Done", blocked="Blocked", progress="In progress"):
    """Green / red / amber on a status column. `row` must be the first row of `rng`."""
    for value, colour in ((done, GREEN), (blocked, RED), (progress, AMBER)):
        if value is None:
            continue
        ws.conditional_formatting.add(rng, FormulaRule(
            formula=[f'${col}{row}="{value}"'],
            fill=PatternFill("solid", fgColor=colour), stopIfTrue=False))


wb = Workbook()

# ============================================================== Start Here tab
ws = wb.active
ws.title = "Start Here"
title_block(ws, "Getting Found in Oakville — six-month tracker",
            "Working file for Luxury Beauty by Cleo R. Companion to seo-roadmap-6-month.docx")

rows = [
    ("", ""),
    ("THE SHAPE OF IT", ""),
    ("Runs", "Monday 17 August 2026 to Sunday 14 February 2027. Twenty-six weeks, six months."),
    ("Right now", "The rest of August is Google Business Profile and citations. Nothing else."),
    ("Week of Aug 24", "Website planning and design. Sitemap, wireframes, keyword-to-page map, copy outline."),
    ("Aug 31", "Design sign-off. This is the date the build waits on."),
    ("Sept 1", "Build starts. Launch target is the week of 5 October."),
    ("", ""),
    ("HOW THIS WORKS", ""),
    ("Every day", "Open the Daily Log. Six slots, five of them daily, the reel on Wednesday. Mark each one."),
    ("The routine", "The Daily Routine tab says what each slot actually means. Read it once, then work off the log."),
    ("The projects", "The Tasks tab holds the one-off work. Filter Status to 'In progress' and 'Not started'."),
    ("Progress", "The 90-Day Overview is where you look at progress. The Tasks tab is where you change it."),
    ("Two citation tabs", "Citations fixes the listings that exist and are wrong. Local Citations builds the ones that do not exist yet."),
    ("You edit", "Status, Date done, and Notes. Leave the rest unless the plan genuinely changes."),
    ("Blocked?", "Set Status to Blocked and write what you are waiting on in Notes. Blocked rows turn red."),
    ("Weekly", "Monday: profile insights. Wednesday: batch the content and film the reel. Friday: log the week."),
    ("Monthly", "Fill the Metrics tab. Compare against the baseline column, not against last month."),
    ("", ""),
    ("THE TWO THINGS THAT DECIDE THIS", ""),
    ("Reviews", "Twelve today. The ask goes out inside 24 hours of every appointment or the number does not move."),
    ("The daily log", "Five slots a day is not a lot. Missing them for a fortnight undoes a month of the rest."),
    ("", ""),
    ("WHERE THINGS STAND", ""),
]
r = 4
for a, b in rows:
    ca = ws.cell(row=r, column=1, value=a)
    cb = ws.cell(row=r, column=2, value=b)
    if a.isupper() and a:
        ca.font = Font(name="Calibri", size=10, bold=True, color="B99753")
    else:
        ca.font = Font(name="Calibri", size=10, bold=True, color=INK)
    cb.font = Font(name="Calibri", size=10, color=INK)
    cb.alignment = Alignment(wrap_text=True, vertical="top")
    r += 1

snap_start = r
n = len(TASKS)
for lbl, formula, fmt in [
    ("Project tasks", f"=COUNTA(Tasks!F2:F{n+1})", None),
    ("Not started", f'=COUNTIF(Tasks!J2:J{n+1},"Not started")', None),
    ("In progress", f'=COUNTIF(Tasks!J2:J{n+1},"In progress")', None),
    ("Blocked", f'=COUNTIF(Tasks!J2:J{n+1},"Blocked")', None),
    ("Done", f'=COUNTIF(Tasks!J2:J{n+1},"Done")', None),
    ("Overdue and not done", f'=SUMPRODUCT((Tasks!D2:D{n+1}<TODAY())*(Tasks!J2:J{n+1}<>"Done")*(Tasks!J2:J{n+1}<>"Not applicable"))', None),
    ("Percent complete", f'=IFERROR(COUNTIF(Tasks!J2:J{n+1},"Done")/COUNTA(Tasks!F2:F{n+1}),0)', "0%"),
    ("Daily slots done", "='Daily Log'!G2", None),
    ("Daily slots missed", "='Daily Log'!I2", None),
    ("Daily completion rate", "='Daily Log'!K2", "0%"),
]:
    ws.cell(row=r, column=1, value=lbl).font = Font(name="Calibri", size=10, bold=True, color=INK)
    c = ws.cell(row=r, column=2, value=formula)
    c.font = Font(name="Calibri", size=10, color=INK)
    if fmt:
        c.number_format = fmt
    r += 1
ws.cell(row=snap_start - 1, column=1).font = Font(name="Calibri", size=10, bold=True, color="B99753")

ws.column_dimensions["A"].width = 26
ws.column_dimensions["B"].width = 96

r += 1
note = ws.cell(row=r, column=1, value="Import to Google Sheets: drive.google.com > New > File upload > pick this file > open with Google Sheets.")
note.font = Font(name="Calibri", size=9, italic=True, color=GREY)

# =========================================================== Daily Routine tab
ws = wb.create_sheet("Daily Routine")
title_block(ws, "The daily routine",
            "Six slots. Five of them every weekday, the reel on Wednesday. This is the tab that "
            "explains the work; the Daily Log is the tab you tick.")
hdr(ws, 4, ["Slot", "Task", "Cadence", "What it involves", "Owner", "Logged on", "Status"],
    [6, 30, 24, 84, 9, 18, 15])

r = 5
routine_ranges = []
for heading, block in (("EVERY DAY", DAILY), ("EVERY WEEK", WEEKLY), ("EVERY MONTH", MONTHLY)):
    band(ws, r, 7, heading)
    r += 1
    first = r
    for slot, task, cadence, detail, owner, where in block:
        for col, v in enumerate([slot, task, cadence, detail, owner, where, "Not started"], start=1):
            c = ws.cell(row=r, column=col, value=v)
            c.font = Font(name="Calibri", size=10, bold=(col == 2), color=INK)
            c.alignment = Alignment(wrap_text=True, vertical="top")
        ws.row_dimensions[r].height = 46
        r += 1
    routine_ranges.append((first, r - 1))

for first, last in routine_ranges:
    dv(ws, STATUSES, f"G{first}:G{last}")
    status_colours(ws, f"A{first}:G{last}", "G", first)
ws.freeze_panes = "A5"

r += 1
for line in [
    "The reel is once a week, on Wednesday, filmed in the Wednesday batch block. Do not try to do one a day.",
    "Before the site is live, the on-page slot is preparation: keyword mapping, title tags, meta descriptions, page copy.",
    "From launch onward it is one real page improved a day, on rotation, so no page goes six months untouched.",
    "Compliance applies to Google posts exactly as it applies to Instagram. Nothing names a prescription drug.",
]:
    c = ws.cell(row=r, column=2, value=line)
    c.font = Font(name="Calibri", size=9, italic=True, color=GREY)
    r += 1

# =============================================================== Daily Log tab
ws = wb.create_sheet("Daily Log")
title_block(ws, "Daily log",
            "One row per weekday, 17 August 2026 to 12 February 2027. Mark each slot Done, Missed "
            "or N/A. The reel column is only live on Wednesdays.")
LOG_COLS = ["Date", "Day", "Week", "GBP post", "IG post / story", "Reel",
            "On-page", "Review ask", "Citation", "Done", "Notes"]
hdr(ws, 4, LOG_COLS, [11, 6, 6, 11, 14, 8, 10, 12, 10, 7, 46])

r = 5
log_rows = []          # every real (non-band) data row
current_month = None
d = START
while d <= END:
    if d.weekday() < 5:                       # weekdays only
        if (d.year, d.month) != current_month:
            current_month = (d.year, d.month)
            band(ws, r, len(LOG_COLS), d.strftime("%B %Y").upper())
            r += 1
        week = (d - START).days // 7 + 1
        is_wed = d.weekday() == 2
        vals = [d, d.strftime("%a"), f"W{week}",
                None, None, (None if is_wed else "N/A"), None, None, None,
                f"=COUNTIF(D{r}:I{r},\"Done\")", None]
        for col, v in enumerate(vals, start=1):
            c = ws.cell(row=r, column=col, value=v)
            c.font = Font(name="Calibri", size=10, color=INK)
            c.alignment = Alignment(vertical="center",
                                    horizontal=("left" if col in (1, 11) else "center"),
                                    wrap_text=(col == 11))
        ws.cell(row=r, column=1).number_format = "yyyy-mm-dd"
        ws.cell(row=r, column=2).font = Font(name="Calibri", size=10, color=GREY)
        ws.cell(row=r, column=3).font = Font(name="Calibri", size=10, color=GREY)
        ws.cell(row=r, column=10).font = Font(name="Calibri", size=10, bold=True, color=INK)
        log_rows.append(r)
        r += 1
    d += dt.timedelta(days=1)

log_first, log_last = log_rows[0], log_rows[-1]
ws.freeze_panes = "A5"

for col in "DEFGHI":
    dv(ws, DAY_MARKS, f"{col}{log_first}:{col}{log_last}")

# A day's target is six on a Wednesday and five on every other weekday. Only
# colour a row once something has actually been marked on it.
marked = f'COUNTIF($D{log_first}:$I{log_first},"Done")+COUNTIF($D{log_first}:$I{log_first},"Missed")>0'
target = f'IF($B{log_first}="Wed",6,5)'
rng = f"A{log_first}:K{log_last}"
ws.conditional_formatting.add(rng, FormulaRule(
    formula=[f'AND({marked},$J{log_first}>={target})'],
    fill=PatternFill("solid", fgColor=GREEN), stopIfTrue=True))
ws.conditional_formatting.add(rng, FormulaRule(
    formula=[f'AND({marked},$J{log_first}={target}-1)'],
    fill=PatternFill("solid", fgColor=AMBER), stopIfTrue=True))
ws.conditional_formatting.add(rng, FormulaRule(
    formula=[f'AND({marked},$J{log_first}<{target}-1)'],
    fill=PatternFill("solid", fgColor=RED), stopIfTrue=True))

summary = [
    (6, "Slots done", 7, f'=COUNTIF(D{log_first}:I{log_last},"Done")', None),
    (8, "Slots missed", 9, f'=COUNTIF(D{log_first}:I{log_last},"Missed")', None),
    (10, "Completion", 11, f'=IFERROR(COUNTIF(D{log_first}:I{log_last},"Done")/'
                           f'(COUNTIF(D{log_first}:I{log_last},"Done")+'
                           f'COUNTIF(D{log_first}:I{log_last},"Missed")),0)', "0%"),
]
for lcol, lbl, vcol, formula, fmt in summary:
    ws.cell(row=2, column=lcol, value=lbl).font = Font(name="Calibri", size=10, bold=True, color=INK)
    c = ws.cell(row=2, column=vcol, value=formula)
    c.font = Font(name="Calibri", size=10, bold=True, color=INK)
    if fmt:
        c.number_format = fmt

counts = [
    (4, "GBP posts", 5, f'=COUNTIF(D{log_first}:D{log_last},"Done")'),
    (6, "Reels", 7, f'=COUNTIF(F{log_first}:F{log_last},"Done")'),
    (8, "Review asks", 9, f'=COUNTIF(H{log_first}:H{log_last},"Done")'),
    (10, "Days logged", 11, f'=COUNTIF(J{log_first}:J{log_last},">0")'),
]
for lcol, lbl, vcol, formula in counts:
    ws.cell(row=3, column=lcol, value=lbl).font = Font(name="Calibri", size=9, bold=True, color=GREY)
    ws.cell(row=3, column=vcol, value=formula).font = Font(name="Calibri", size=9, color=GREY)

# ========================================================= 90-Day Overview tab
ws = wb.create_sheet("90-Day Overview")
title_block(ws, "The first 90 days",
            "17 August to 15 November 2026. Week 13 closes day 90. Counts pull live from the "
            "Tasks tab — this is where you look at progress, that is where you change it.")

OV_COLS = 9
band(ws, 4, OV_COLS, "THE THREE CHECKPOINTS — what has to be true, and by when")
hdr(ws, 5, ["Checkpoint", "Date", "What must be true by then", "", "", "",
            "On track", "Notes", ""],
    [13, 12, 30, 14, 14, 12, 13, 30, 20])
r = 6
cp_first = r
for name, when, must in CHECKPOINTS:
    ws.merge_cells(start_row=r, start_column=3, end_row=r, end_column=6)
    for col, v in enumerate([name, when, must, None, None, None, None, None, None], start=1):
        c = ws.cell(row=r, column=col, value=v)
        c.font = Font(name="Calibri", size=10, bold=(col == 1), color=INK)
        c.alignment = Alignment(vertical="top", wrap_text=(col in (3, 8)))
    ws.cell(row=r, column=2).number_format = "yyyy-mm-dd"
    ws.row_dimensions[r].height = 46
    r += 1
cp_last = r - 1
dv(ws, ["Yes", "At risk", "No"], f"G{cp_first}:G{cp_last}")
status_colours(ws, f"A{cp_first}:I{cp_last}", "G", cp_first,
               done="Yes", blocked="No", progress="At risk")

r += 1
band(ws, r, OV_COLS, "WEEK BY WEEK — the first 90 days")
r += 1
hdr(ws, r, ["Week", "Dates", "Phase", "The one thing that matters",
            "Tasks due", "Done", "Progress", "Status", "Notes"],
    [7, 20, 20, 62, 11, 8, 10, 14, 30])
r += 1
wk_first = r
for week, phase, matters in QUARTER_WEEKS:
    start = START + dt.timedelta(weeks=week - 1)
    end = WEEK_DUE[week]
    dates = f"{start:%d %b} – {end:%d %b}"
    vals = [f"W{week}", dates, phase, matters,
            f'=COUNTIF(Tasks!$C:$C,"W{week}")',
            f'=COUNTIFS(Tasks!$C:$C,"W{week}",Tasks!$J:$J,"Done")',
            f"=IFERROR(F{r}/E{r},0)", "Not started", None]
    for col, v in enumerate(vals, start=1):
        c = ws.cell(row=r, column=col, value=v)
        c.font = Font(name="Calibri", size=10, bold=(col == 1), color=INK)
        c.alignment = Alignment(vertical="top",
                                horizontal=("center" if col in (1, 5, 6, 7) else "left"),
                                wrap_text=(col in (4, 9)))
    ws.cell(row=r, column=7).number_format = "0%"
    ws.row_dimensions[r].height = 28
    r += 1
wk_last = r - 1

dv(ws, STATUSES, f"H{wk_first}:H{wk_last}")
wk_rng = f"A{wk_first}:I{wk_last}"
# Green once every task for the week is done; amber part-way; red only once the
# week has actually ended with nothing closed out.
ws.conditional_formatting.add(wk_rng, FormulaRule(
    formula=[f'AND($E{wk_first}>0,$F{wk_first}=$E{wk_first})'],
    fill=PatternFill("solid", fgColor=GREEN), stopIfTrue=True))
ws.conditional_formatting.add(wk_rng, FormulaRule(
    formula=[f'$F{wk_first}>0'],
    fill=PatternFill("solid", fgColor=AMBER), stopIfTrue=True))
# A week only goes red once it has actually ended with nothing closed out, so
# one rule per row carrying that row's own end date.
for i, (week, _, _) in enumerate(QUARTER_WEEKS):
    row = wk_first + i
    due = WEEK_DUE[week]
    ws.conditional_formatting.add(f"A{row}:I{row}", FormulaRule(
        formula=[f'AND($E{row}>0,$F{row}=0,TODAY()>DATE({due.year},{due.month},{due.day}))'],
        fill=PatternFill("solid", fgColor=RED), stopIfTrue=True))

r += 1
band(ws, r, OV_COLS, "AFTER DAY 90 — the second half")
r += 1
hdr(ws, r, ["Month", "Dates", "Theme", "", "", "", "", "", ""],
    [7, 20, 62, 11, 8, 10, 14, 30, 20])
r += 1
for m, dates, theme in [
    ("Month 4", "9 Nov – 6 Dec", "Depth on the pages that earn. Thin pages expanded, gallery, articles four to six"),
    ("Month 5", "7 Dec – 10 Jan", "Hold the cadence through the holidays. Reduced posting, never stopped posting"),
    ("Month 6", "11 Jan – 14 Feb", "Compound it. Near-miss queries, technical audit, the six-month report"),
]:
    ws.merge_cells(start_row=r, start_column=3, end_row=r, end_column=8)
    for col, v in enumerate([m, dates, theme], start=1):
        c = ws.cell(row=r, column=col, value=v)
        c.font = Font(name="Calibri", size=10, bold=(col == 1), color=INK)
        c.alignment = Alignment(vertical="top", wrap_text=(col == 3))
    ws.row_dimensions[r].height = 26
    r += 1

c = ws.cell(row=r, column=1, value="Week 13 is both the day-90 review and the first week of month four. "
                                   "It appears in two places on purpose.")
c.font = Font(name="Calibri", size=9, italic=True, color=GREY)
ws.merge_cells(start_row=r, start_column=1, end_row=r, end_column=OV_COLS)
r += 2

band(ws, r, OV_COLS, "THE DATES THAT DO NOT MOVE")
r += 1
hdr(ws, r, ["Date", "What", "Why it matters", "", "", "", "Status", "Notes", ""],
    [12, 22, 58, 11, 8, 10, 14, 30, 20])
r += 1
hd_first = r
for when, what, why in HARD_DATES:
    ws.merge_cells(start_row=r, start_column=3, end_row=r, end_column=6)
    for col, v in enumerate([when, what, why, None, None, None, "Not started", None, None], start=1):
        c = ws.cell(row=r, column=col, value=v)
        c.font = Font(name="Calibri", size=10, bold=(col == 2), color=INK)
        c.alignment = Alignment(vertical="top", wrap_text=(col in (3, 8)))
    ws.cell(row=r, column=1).number_format = "yyyy-mm-dd"
    r += 1
hd_last = r - 1
dv(ws, STATUSES, f"G{hd_first}:G{hd_last}")
status_colours(ws, f"A{hd_first}:I{hd_last}", "G", hd_first)
ws.conditional_formatting.add(f"A{hd_first}:I{hd_last}", FormulaRule(
    formula=[f'AND($A{hd_first}<TODAY(),$G{hd_first}<>"Done")'],
    font=Font(color="B03A2E", bold=True), stopIfTrue=False))

ws.freeze_panes = "A4"

# =================================================================== Tasks tab
ws = wb.create_sheet("Tasks")
headers = ["ID", "Month", "Week", "Due", "Area", "Task", "Owner",
           "Cadence", "Priority", "Status", "Date done", "Notes / blockers"]
widths = [5, 9, 7, 11, 15, 74, 9, 13, 9, 14, 11, 40]
hdr(ws, 1, headers, widths)

for i, (week, area, task, owner, cadence, prio) in enumerate(TASKS, start=1):
    row = i + 1
    vals = [i, f"Month {MONTH_OF_WEEK[week]}", f"W{week}", WEEK_DUE[week], area, task,
            owner, cadence, prio, "Not started", None, None]
    for col, v in enumerate(vals, start=1):
        c = ws.cell(row=row, column=col, value=v)
        c.font = Font(name="Calibri", size=10, color=INK)
        c.alignment = Alignment(vertical="top", wrap_text=(col in (6, 12)))
        if col in (4, 11):
            c.number_format = "yyyy-mm-dd"
    ws.row_dimensions[row].height = 15

last = len(TASKS) + 1
ws.freeze_panes = "A2"
ws.auto_filter.ref = f"A1:L{last}"

dv(ws, OWNERS, f"G2:G{last}")
dv(ws, CADENCES, f"H2:H{last}")
dv(ws, PRIORITIES, f"I2:I{last}")
dv(ws, STATUSES, f"J2:J{last}")

status_colours(ws, f"A2:L{last}", "J", 2)
ws.conditional_formatting.add(f"A2:L{last}", FormulaRule(
    formula=['AND($D2<TODAY(),$J2<>"Done",$J2<>"Not applicable")'],
    font=Font(color="B03A2E", bold=True), stopIfTrue=False))

# ============================================================== Site Build tab
ws = wb.create_sheet("Site Build")
title_block(ws, "Website build timeline",
            "Approved, full build. Real dates, not weeks-from-approval. Everything after "
            "31 August slips one-for-one if the design sign-off slips.")
hdr(ws, 4, ["When", "What we do", "What we need from Cleo", "Status", "Date done", "Notes"],
    [19, 62, 44, 14, 11, 30])
build = [
    ("Aug 24 - 28", "Planning and design: sitemap, page inventory, wireframes for home, contact and the "
                    "treatment template, keyword-to-page map, brand kit pulled from design/brand",
                    "Registrar login, treatment list with current pricing, photos of the room, a headshot, logo files"),
    ("Aug 31", "Design sign-off. This is the gate — nothing is built before it clears",
               "Sign off the wireframes and the look, in writing"),
    ("Sep 1 - 6", "BUILD STARTS. Hosting sorted, DNS pointed at the new site, platform set up, "
                  "page structure scaffolded", "Nothing. This week is ours"),
    ("Sep 7 - 13", "Home and contact complete in staging: corrected address at L6M 0T9, real hours, "
                   "booking link, LocalBusiness and Service schema, title tags and meta descriptions",
                   "Read the home page copy and approve it"),
    ("Sep 14 - 27", "Treatment pages: lip filler, dermal filler and facial balancing, anti-wrinkle "
                    "injections, skin boosters. Internal linking, image compression, alt text",
                    "Read the copy and confirm every claim is one you can stand behind"),
    ("Sep 28 - Oct 4", "Remaining treatment pages: dermaplaning, electrolysis, Belkyra. The Oakville "
                       "and surrounding-area page. Mobile speed pass", "Nothing. This week is ours"),
    ("Oct 5 - 11", "LAUNCH. Search Console verified, sitemap submitted, analytics running, GBP website "
                   "field repointed, every citation updated to the new URL", "Final sign-off before it goes live"),
    ("Oct 12 - Nov 8", "Indexing and on-page. Fix whatever Search Console flags, FAQ schema on the "
                       "treatment pages, the first three question-intent articles",
                       "The questions you get asked most in consults"),
    ("Nov 9 - Feb 14", "Depth. Weight-loss page, before-and-after gallery, five more articles, "
                       "near-miss query optimisation, technical audit, handover",
                       "Consent-cleared photos as they come, and the consult questions"),
]
for i, (when, doing, needs) in enumerate(build, start=5):
    for col, v in enumerate([when, doing, needs, "Not started", None, None], start=1):
        c = ws.cell(row=i, column=col, value=v)
        c.font = Font(name="Calibri", size=10, bold=(col == 1), color=INK)
        c.alignment = Alignment(wrap_text=True, vertical="top")
    ws.cell(row=i, column=5).number_format = "yyyy-mm-dd"
    ws.row_dimensions[i].height = 42
last_b = 4 + len(build)
dv(ws, STATUSES, f"D5:D{last_b}")
status_colours(ws, f"A5:F{last_b}", "D", 5)
ws.freeze_panes = "A5"

# ================================================================= Reviews tab
ws = wb.create_sheet("Reviews")
title_block(ws, "Review tracker",
            "Baseline 12 reviews at 5.0. Target 18 to 22 new ones across the six months, "
            "roughly one every ten days. Ask inside 24 hours or the number does not move.")
hdr(ws, 4, ["Date asked", "Patient (initials)", "How we asked", "Asked by",
            "Review received", "Date received", "Platform", "Replied", "Notes"],
    [12, 17, 16, 11, 15, 13, 12, 10, 40])
REV_LAST = 164
for row in range(5, REV_LAST + 1):
    for col in range(1, 10):
        ws.cell(row=row, column=col).font = Font(name="Calibri", size=10, color=INK)
    ws.cell(row=row, column=1).number_format = "yyyy-mm-dd"
    ws.cell(row=row, column=6).number_format = "yyyy-mm-dd"
ws.freeze_panes = "A5"
dv(ws, ["WhatsApp", "QR card in room", "In person", "Email", "Text"], f"C5:C{REV_LAST}")
dv(ws, ["Cleo", "Agency"], f"D5:D{REV_LAST}")
dv(ws, ["Yes", "No", "Waiting"], f"E5:E{REV_LAST}")
dv(ws, ["Google", "Facebook", "Birdeye", "Fresha"], f"G5:G{REV_LAST}")
dv(ws, ["Yes", "No"], f"H5:H{REV_LAST}")
ws.conditional_formatting.add(f"A5:I{REV_LAST}", FormulaRule(
    formula=['$E5="Yes"'], fill=PatternFill("solid", fgColor=GREEN), stopIfTrue=False))

for col, lbl, formula in [(6, "Asks sent", f"=COUNTA(A5:A{REV_LAST})"),
                          (8, "Reviews landed", f'=COUNTIF(E5:E{REV_LAST},"Yes")')]:
    ws.cell(row=2, column=col, value=lbl).font = Font(name="Calibri", size=10, bold=True, color=INK)
    ws.cell(row=2, column=col + 1, value=formula).font = Font(name="Calibri", size=10, bold=True, color=INK)

# =============================================================== Citations tab
ws = wb.create_sheet("Citations")
title_block(ws, "Citations and NAP",
            "One name, one address, one phone number, everywhere. The postal code is L6M 0T9 — "
            "it is wrong on most of these today.")
hdr(ws, 4, ["Directory", "Listing URL", "Claimed", "Name correct", "Address and postal code correct",
            "Phone correct", "Website URL current", "Status", "Date done", "Notes"],
    [26, 34, 11, 13, 30, 13, 18, 14, 11, 34])
directories = [
    "Google Business Profile", "Birdeye", "Facebook", "Fresha", "Yelp.ca", "Yellow Pages",
    "Apple Maps", "Bing Places", "RateMDs", "Canadian chamber directories",
    "Oakville Chamber of Commerce", "411.ca", "Cylex Canada", "Foursquare",
    "Instagram profile", "Booking system profile", "Halton business directory",
]
CIT_LAST = 4 + len(directories) + 13      # blank rows for whatever the audit turns up
for i in range(5, CIT_LAST + 1):
    name = directories[i - 5] if i - 5 < len(directories) else None
    for col, v in enumerate([name, None, None, None, None, None, None,
                             ("Not started" if name else None), None, None], start=1):
        c = ws.cell(row=i, column=col, value=v)
        c.font = Font(name="Calibri", size=10, bold=(col == 1), color=INK)
        c.alignment = Alignment(wrap_text=True, vertical="top")
    ws.cell(row=i, column=9).number_format = "yyyy-mm-dd"
for col in "CDEFG":
    dv(ws, ["Yes", "No", "N/A"], f"{col}5:{col}{CIT_LAST}")
dv(ws, STATUSES, f"H5:H{CIT_LAST}")
status_colours(ws, f"A5:J{CIT_LAST}", "H", 5)
ws.freeze_panes = "A5"
ws.auto_filter.ref = f"A4:J{CIT_LAST}"

# ========================================================= Local Citations tab
ws = wb.create_sheet("Local Citations")
title_block(ws, "Local citations — the build list",
            "The Citations tab fixes listings that already exist. This one builds listings "
            "where there are none. Decide the record once, then paste it, unchanged, into "
            "every form below.")

band(ws, 4, 11, "THE RECORD — paste this, exactly, every time")
for col, h in enumerate(["Field", "Value", "Where it gets decided"], start=1):
    c = ws.cell(row=5, column=(1 if col == 1 else (2 if col == 2 else 5)), value=h)
    c.font = Font(name="Calibri", size=10, bold=True, color=INK)
    c.fill = PatternFill("solid", fgColor=CREAM)
    c.border = Border(bottom=Side("thin", color=GOLD))

r = 6
rec_first = r
for field, value, note in RECORD:
    ws.merge_cells(start_row=r, start_column=2, end_row=r, end_column=4)
    ws.merge_cells(start_row=r, start_column=5, end_row=r, end_column=11)
    a = ws.cell(row=r, column=1, value=field)
    a.font = Font(name="Calibri", size=10, bold=True, color=INK)
    a.alignment = Alignment(vertical="center", wrap_text=True)
    b = ws.cell(row=r, column=2, value=(value or None))
    b.font = Font(name="Calibri", size=10, color=INK)
    b.alignment = Alignment(vertical="center", wrap_text=True)
    n = ws.cell(row=r, column=5, value=(note or None))
    n.font = Font(name="Calibri", size=9, italic=True, color=GREY)
    n.alignment = Alignment(vertical="center", wrap_text=True)
    r += 1
rec_last = r - 1

# Anything still undecided glows amber until someone fills it in.
ws.conditional_formatting.add(f"B{rec_first}:B{rec_last}", FormulaRule(
    formula=[f'$B{rec_first}=""'], fill=PatternFill("solid", fgColor=AMBER), stopIfTrue=False))
c = ws.cell(row=r, column=1, value="Amber means still undecided. Every amber row above has a matching "
                                   "row on the Waiting On Cleo tab. Do not submit to anything below "
                                   "until they are all filled in — a wrong detail submitted 28 times "
                                   "is 28 corrections later.")
c.font = Font(name="Calibri", size=9, italic=True, color=GREY)
c.alignment = Alignment(wrap_text=True, vertical="top")
ws.merge_cells(start_row=r, start_column=1, end_row=r, end_column=11)
ws.row_dimensions[r].height = 26
r += 2

band(ws, r, 11, "SUBMISSION QUEUE — one a day in the citation slot, once the record is settled")
r += 1
hdr(ws, r, ["Directory", "Type", "Priority", "Where to start", "Account needed",
            "Verified live", "Status", "Date submitted", "Live listing URL",
            "NAP checked", "Notes"],
    [32, 13, 10, 24, 15, 13, 14, 14, 32, 13, 34])
q_hdr = r
r += 1
q_first = r
for directory, ctype, prio, where in LOCAL_CITATIONS:
    for col, v in enumerate([directory, ctype, prio, (where or None), None, None,
                             "Not started", None, None, None, None], start=1):
        c = ws.cell(row=r, column=col, value=v)
        c.font = Font(name="Calibri", size=10, bold=(col == 1), color=INK)
        c.alignment = Alignment(vertical="top", wrap_text=(col in (1, 4, 9, 11)))
    ws.cell(row=r, column=8).number_format = "yyyy-mm-dd"
    r += 1
q_last = r - 1

dv(ws, CITATION_TYPES, f"B{q_first}:B{q_last}")
dv(ws, PRIORITIES, f"C{q_first}:C{q_last}")
for col in "EFJ":
    dv(ws, ["Yes", "No", "N/A"], f"{col}{q_first}:{col}{q_last}")
dv(ws, STATUSES, f"G{q_first}:G{q_last}")
status_colours(ws, f"A{q_first}:K{q_last}", "G", q_first)
ws.auto_filter.ref = f"A{q_hdr}:K{q_last}"

for col, lbl, formula in [(6, "Submitted", f'=COUNTIF(G{q_first}:G{q_last},"Done")'),
                          (8, "Still to do", f'=COUNTIF(G{q_first}:G{q_last},"Not started")')]:
    ws.cell(row=2, column=col, value=lbl).font = Font(name="Calibri", size=10, bold=True, color=INK)
    ws.cell(row=2, column=col + 1, value=formula).font = Font(name="Calibri", size=10, bold=True, color=INK)

r = q_last + 2
for line in [
    "Domains in 'Where to start' are a starting point, not gospel. Canadian directories close and change hands. "
    "The first pass on every row is confirming it still exists and still takes a free listing — if it does not, "
    "set Status to Not applicable and say so in Notes rather than leaving it silent.",
    "Google, Yelp.ca, Facebook, Fresha, Yellow Pages, Apple Maps, Bing Places, RateMDs, 411.ca, Cylex, Foursquare "
    "and the chamber listings are not repeated here — they already exist and live on the Citations tab.",
    "Every listing built before the site launches carries the booking link. In launch week, week of 5 October, "
    "every one of them gets updated to the live site URL. That is a task on the Tasks tab, not a thing to remember.",
]:
    c = ws.cell(row=r, column=1, value=line)
    c.font = Font(name="Calibri", size=9, italic=True, color=GREY)
    c.alignment = Alignment(wrap_text=True, vertical="top")
    ws.merge_cells(start_row=r, start_column=1, end_row=r, end_column=11)
    ws.row_dimensions[r].height = 26
    r += 1

# =============================================================== GBP Posts tab
ws = wb.create_sheet("GBP Posts")
title_block(ws, "Google Business Profile posts",
            "One a day, every weekday. Rotate the five types. Every post carries the booking link. "
            "Nothing here names a prescription drug.")
hdr(ws, 4, ["Date", "Post type", "Treatment or topic", "Copy source", "Image source",
            "Booking link included", "Published", "Notes"],
    [11, 24, 30, 26, 26, 19, 12, 34])
GBP_LAST = 4 + 140
for i in range(5, GBP_LAST + 1):
    for col in range(1, 9):
        ws.cell(row=i, column=col).font = Font(name="Calibri", size=10, color=INK)
    ws.cell(row=i, column=1).number_format = "yyyy-mm-dd"
dv(ws, ["Treatment spotlight", "FAQ or myth", "Before and after",
        "Offer or availability", "Practice update", "Reel or video"], f"B5:B{GBP_LAST}")
dv(ws, ["Yes", "No"], f"F5:F{GBP_LAST}")
dv(ws, ["Yes", "No", "Scheduled"], f"G5:G{GBP_LAST}")
ws.conditional_formatting.add(f"A5:H{GBP_LAST}", FormulaRule(
    formula=['$G5="Yes"'], fill=PatternFill("solid", fgColor=GREEN), stopIfTrue=False))
ws.conditional_formatting.add(f"A5:H{GBP_LAST}", FormulaRule(
    formula=['$G5="Scheduled"'], fill=PatternFill("solid", fgColor=AMBER), stopIfTrue=False))
ws.freeze_panes = "A5"
ws.cell(row=2, column=6, value="Posts published").font = Font(name="Calibri", size=10, bold=True, color=INK)
ws.cell(row=2, column=7, value=f'=COUNTIF(G5:G{GBP_LAST},"Yes")').font = Font(name="Calibri", size=10, bold=True, color=INK)

# ================================================================= Metrics tab
ws = wb.create_sheet("Metrics")
title_block(ws, "Monthly numbers",
            "Fill at the end of each month. Compare against the baseline column, not against last month.")
hdr(ws, 4, ["What we track", "Baseline (17 Aug)", "Target by 14 Feb",
            "End Sep", "End Oct", "End Nov", "End Dec", "End Jan", "Mid Feb", "Notes"],
    [38, 17, 34, 10, 10, 10, 10, 10, 10, 34])
metrics = [
    ("Google reviews", "12", "30 to 34, still at or near 5.0"),
    ("Google review average", "5.0", "At or near 5.0"),
    ("Calls from Google profile", "", "Up on baseline every month"),
    ("Direction requests from Google", "", "Up on baseline every month"),
    ("Google profile views", "", "Up on baseline every month"),
    ("GBP posts published in the month", "0", "20 or more, every month"),
    ("Map pack position: nurse injector oakville", "", "Visible in the pack"),
    ("Map pack position: botox oakville", "", "Visible in the pack"),
    ("Map pack position: lip filler oakville", "", "Visible in the pack"),
    ("Instagram profile visits", "", "Up on baseline"),
    ("Instagram link taps", "", "Rising faster than profile visits"),
    ("Reels published in the month", "0", "Four, one a week"),
    ("DMs received", "Not tracked", "Tracked weekly from week 1"),
    ("Consultations booked", "Not tracked", "Tracked weekly, attributed to source"),
    ("Ad spend", "Not supplied", "Confirmed and split 70/20/10"),
    ("Cost per DM from ads", "", "Falling month on month"),
    ("Website sessions", "No site", "From launch in October, rising month on month"),
    ("Search Console impressions", "No site", "Rising month on month from October"),
    ("Pages indexed", "No site", "Every page, confirmed in Search Console"),
    ("Daily slot completion rate", "", "85 per cent or better"),
]
for i, (m, base, target) in enumerate(metrics, start=5):
    for col, v in enumerate([m, base, target], start=1):
        c = ws.cell(row=i, column=col, value=v)
        c.font = Font(name="Calibri", size=10, bold=(col == 1), color=INK)
        c.alignment = Alignment(wrap_text=True, vertical="top")
ws.freeze_panes = "D5"

# ============================================================== Waiting On tab
ws = wb.create_sheet("Waiting On Cleo")
title_block(ws, "Waiting on Cleo",
            "Access, assets and answers. Most delay in work like this comes from waiting on these. "
            "The top four are what the build waits on.")
hdr(ws, 4, ["Type", "What we need", "Needed by", "Why it blocks us", "Status", "Date received", "Notes"],
    [11, 46, 13, 46, 14, 13, 28])
waiting = [
    ("Access", "Google Business Profile, and which email holds it", dt.date(2026, 8, 21),
     "Nothing on the profile can start until this is settled"),
    ("Answer", "Real opening hours for every day", dt.date(2026, 8, 21),
     "The hours on Google are wrong until we have these"),
    ("Asset", "Treatment list with current pricing", dt.date(2026, 8, 26),
     "Sitemap, service pages and the GBP services list all key off it"),
    ("Access", "Domain registrar login for luxurybeautybycleo.ca", dt.date(2026, 8, 28),
     "DNS has to be pointed before the build can start on 1 September"),
    ("Answer", "DESIGN SIGN-OFF on the wireframes and the look", dt.date(2026, 8, 31),
     "The build gate. Everything after it slips one-for-one"),
    ("Asset", "Photos of the treatment room and the space", dt.date(2026, 8, 28),
     "GBP photos and every website page"),
    ("Asset", "A current headshot", dt.date(2026, 8, 28),
     "GBP, Instagram and the site"),
    ("Asset", "Logo files and any brand assets you already have", dt.date(2026, 8, 28),
     "The design cannot be finalised without them"),
    ("Asset", "Signed before-and-after consent form", dt.date(2026, 9, 6),
     "Determines what we are allowed to publish anywhere"),
    ("Access", "Meta Business account and ad account", dt.date(2026, 9, 20),
     "Needed to build audiences and run ads"),
    ("Access", "Instagram collaborator invitation", dt.date(2026, 8, 28),
     "Needed to fix the name field, bio and link"),
    ("Access", "Booking system login", dt.date(2026, 9, 6),
     "Needed for the booking link and the lookalike audience"),
    ("Answer", "Whether Allure allows your own signage and suite designation", dt.date(2026, 8, 23),
     "Decides how we build the Google listing"),
    ("Answer", "The three competitors you want to be measured against", dt.date(2026, 8, 23),
     "Shapes what we track and benchmark"),
    ("Answer", "Your ideal patient, in your own words", dt.date(2026, 9, 13),
     "Ad audiences and page copy"),
    ("Answer", "Current monthly spend on Instagram and Google ads", dt.date(2026, 9, 13),
     "Cannot recommend a budget without it"),
    ("Answer", "Approval of the home page copy", dt.date(2026, 9, 13),
     "Home and contact cannot go into staging without it"),
    ("Answer", "Approval of the treatment page copy", dt.date(2026, 9, 27),
     "Every claim has to be one you can stand behind"),
    ("Answer", "The questions you get asked most in consults", dt.date(2026, 10, 18),
     "The articles are written from them, not invented"),
    ("Answer", "Commitment to the daily routine", dt.date(2026, 8, 21),
     "Five slots a day. Without them the rest of the plan does not compound"),
]
for i, (t, need, by, why) in enumerate(waiting, start=5):
    for col, v in enumerate([t, need, by, why, "Not started", None, None], start=1):
        c = ws.cell(row=i, column=col, value=v)
        c.font = Font(name="Calibri", size=10, bold=(col == 1), color=INK)
        c.alignment = Alignment(wrap_text=True, vertical="top")
    ws.cell(row=i, column=3).number_format = "yyyy-mm-dd"
    ws.cell(row=i, column=6).number_format = "yyyy-mm-dd"
last_w = 4 + len(waiting)
dv(ws, ["Not started", "Asked", "Chased", "Received", "Not applicable"], f"E5:E{last_w}")
status_colours(ws, f"A5:G{last_w}", "E", 5, done="Received", blocked=None, progress="Chased")
ws.conditional_formatting.add(f"A5:G{last_w}", FormulaRule(
    formula=['AND($C5<TODAY(),$E5<>"Received",$E5<>"Not applicable")'],
    font=Font(color="B03A2E", bold=True), stopIfTrue=False))
ws.freeze_panes = "A5"

wb.save(OUT)
print(f"wrote {OUT}")
print(f"  {len(TASKS)} project tasks across {WEEKS} weeks, "
      f"{WEEK_DUE[1]:%Y-%m-%d} to {WEEK_DUE[WEEKS]:%Y-%m-%d}")
print(f"  {len(log_rows)} weekday rows in the daily log, "
      f"{len(DAILY)} daily slots, {len(WEEKLY)} weekly, {len(MONTHLY)} monthly")
print(f"  {len(wb.sheetnames)} tabs: {', '.join(wb.sheetnames)}")
